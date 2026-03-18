"""
make_table1.py
Patient characteristics: Overall + stratified by hospice enrollment
Columns: Overall | Hospice | No Hospice | SMD
Output: C:/Users/hsaee/Desktop/CMS_viewer/projects/HNC_io_hosp/table1.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\table1.xlsx"

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading io_analytic...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        hospice_enrolled,
        age_at_death, age_cat,
        sex, race,
        dual_eligible,
        census_region, urban_rural,
        subsite_category,
        io_agent, io_regimen,
        last_episode_doses,
        primary_curative_type,
        had_surgery, had_radiation,
        van_walraven_score,
        days_last_io_to_death,
        in_hospital_death,
        death_year
    FROM io_analytic
""").df()
con.close()

df['hospice_enrolled'] = df['hospice_enrolled'].fillna(0).astype(int)
df['in_hospital_death'] = df['in_hospital_death'].fillna(0).astype(int)
df['dual_eligible'] = df['dual_eligible'].fillna(0).astype(int)
df['had_surgery'] = df['had_surgery'].fillna(0).astype(int)
df['had_radiation'] = df['had_radiation'].fillna(0).astype(int)

hosp   = df[df['hospice_enrolled'] == 1]
nohosp = df[df['hospice_enrolled'] == 0]

print(f"  Total: {len(df):,}  Hospice: {len(hosp):,}  No hospice: {len(nohosp):,}")

# ── SMD helpers ───────────────────────────────────────────────────────────────
def smd_bin(col, val=None):
    """SMD for binary: |p1 - p2| / sqrt((p1(1-p1) + p2(1-p2)) / 2)"""
    p1 = (hosp[col] == val).mean() if val is not None else hosp[col].mean()
    p2 = (nohosp[col] == val).mean() if val is not None else nohosp[col].mean()
    denom = np.sqrt((p1*(1-p1) + p2*(1-p2)) / 2)
    return float(abs(p1 - p2) / denom) if denom > 0 else 0.0

def smd_cont(col):
    """SMD for continuous: |mean1 - mean2| / sqrt((var1 + var2) / 2)"""
    a = hosp[col].dropna()
    b = nohosp[col].dropna()
    denom = np.sqrt((a.var() + b.var()) / 2)
    return float(abs(a.mean() - b.mean()) / denom) if denom > 0 else 0.0

def smd_cat(col):
    """Categorical SMD: max binary SMD across all levels"""
    cats = df[col].dropna().unique()
    return max((smd_bin(col, c) for c in cats), default=0.0)

def fmt_smd(s):
    return f"{s:.2f}"

# ── Format helpers ─────────────────────────────────────────────────────────────
def n_pct(series, val=None):
    if val is not None:
        n = (series == val).sum()
        p = 100.0 * n / len(series)
    else:
        n = series.sum()
        p = 100.0 * n / len(series)
    return f"{n:,} ({p:.1f}%)"

def med_iqr(series):
    s = series.dropna()
    return f"{s.median():.0f} ({s.quantile(0.25):.0f}–{s.quantile(0.75):.0f})"

def mean_sd(series):
    s = series.dropna()
    return f"{s.mean():.1f} ({s.std():.1f})"

# ── Build rows ────────────────────────────────────────────────────────────────
rows = []

def section(title):
    rows.append({'label': title, 'is_section': True,
                 'overall': '', 'hospice': '', 'no_hospice': '', 'smd': ''})

def add_cont(label, col, fmt='median', indent=False):
    fn = med_iqr if fmt == 'median' else mean_sd
    try:
        smd = fmt_smd(smd_cont(col))
    except Exception:
        smd = ''
    rows.append({
        'label':      ('    ' + label) if indent else label,
        'is_section': False,
        'overall':    fn(df[col]),
        'hospice':    fn(hosp[col]),
        'no_hospice': fn(nohosp[col]),
        'smd':        smd,
    })

def add_bin(label, col, val=None, show_smd=True, indent=False):
    try:
        smd = fmt_smd(smd_bin(col, val)) if show_smd else ''
    except Exception:
        smd = ''
    rows.append({
        'label':      ('    ' + label) if indent else label,
        'is_section': False,
        'overall':    n_pct(df[col], val),
        'hospice':    n_pct(hosp[col], val),
        'no_hospice': n_pct(nohosp[col], val),
        'smd':        smd,
    })

def add_cat(label, col, indent=False):
    """Add SMD for overall category, then sub-rows for each level."""
    try:
        smd = fmt_smd(smd_cat(col))
    except Exception:
        smd = ''
    rows.append({
        'label':      ('    ' + label) if indent else label,
        'is_section': False,
        'overall':    '', 'hospice': '', 'no_hospice': '', 'smd': smd,
    })
    for val in sorted(df[col].dropna().unique()):
        rows.append({
            'label': '        ' + str(val),
            'is_section': False,
            'overall':    n_pct(df[col], val),
            'hospice':    n_pct(hosp[col], val),
            'no_hospice': n_pct(nohosp[col], val),
            'smd': '',
        })

# ── Demographics ───────────────────────────────────────────────────────────────
section('DEMOGRAPHICS')
add_cont('Age at death, median (IQR)', 'age_at_death')
add_bin('Male sex', 'sex', 'Male')
add_cat('Race/ethnicity', 'race')
add_bin('Dual eligible (Medicaid)', 'dual_eligible', 1)
add_cat('Urban/rural', 'urban_rural')

section('CLINICAL CHARACTERISTICS')
add_cat('HNC subsite', 'subsite_category')
add_cat('IO agent', 'io_agent')
add_cat('IO regimen', 'io_regimen')
add_cont('Last episode IO doses, median (IQR)', 'last_episode_doses')

section('PRIOR CURATIVE THERAPY')
add_cat('Primary curative type', 'primary_curative_type')

section('COMORBIDITY')
add_cont('van Walraven score, mean (SD)', 'van_walraven_score', fmt='mean')

section('OUTCOMES')
add_bin('Hospice enrolled', 'hospice_enrolled', 1)
add_bin('In-hospital death', 'in_hospital_death', 1)
add_cont('Days from last IO to death, median (IQR)', 'days_last_io_to_death')

# ── Build DataFrame ────────────────────────────────────────────────────────────
df_table = pd.DataFrame(rows)

# ── Write Excel ────────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='BA0C2F')   # OSU Scarlet
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='F5D0D6')   # Light scarlet tint
SECTION_FONT = Font(bold=True, size=10, color='7A0820')
ALT_FILL     = PatternFill('solid', fgColor='F9ECEE')   # Very light tint
TITLE_FONT   = Font(bold=True, size=13, color='BA0C2F')
SMD_HIGH     = PatternFill('solid', fgColor='FFE699')   # Highlight SMD >=0.20

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Table 1'

ws.append([f'Table 1. Patient Characteristics (N = {len(df):,})'])
ws['A1'].font = TITLE_FONT
ws.append([])

n_hosp   = len(hosp)
n_nohosp = len(nohosp)

# Headers
header_row = ws.max_row + 1
cols = [
    'Characteristic',
    f'Overall\n(N = {len(df):,})',
    f'Hospice Enrolled\n(n = {n_hosp:,})',
    f'No Hospice\n(n = {n_nohosp:,})',
    'SMD',
]
for ci, cn in enumerate(cols, 1):
    cell = ws.cell(row=header_row, column=ci, value=cn)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                wrap_text=True, vertical='center')
ws.row_dimensions[header_row].height = 36

alt = 0
for ri, row_data in enumerate(df_table.itertuples(index=False), header_row + 1):
    is_sec = row_data.is_section
    if not is_sec:
        alt += 1
    vals = [row_data.label, row_data.overall, row_data.hospice,
            row_data.no_hospice, row_data.smd]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if is_sec:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif alt % 2 == 0:
            cell.fill = ALT_FILL
        # Highlight SMD >= 0.20 (potentially important imbalance)
        if ci == 5 and not is_sec and val not in ('', None):
            try:
                if float(val) >= 0.20:
                    cell.fill = SMD_HIGH
            except (ValueError, TypeError):
                pass
        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center', wrap_text=(ci == 1))

ws.column_dimensions['A'].width = 44
for letter in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[letter].width = 22

ws.freeze_panes = f'B{header_row + 1}'

footer_row = ws.max_row + 2
ws.cell(row=footer_row, column=1,
        value='Continuous variables: median (IQR); comorbidity: mean (SD). '
              'Binary/categorical: n (%). '
              'SMD = standardized mean difference; '
              '<0.10 negligible, 0.10–<0.20 modest, ≥0.20 potentially important (highlighted).')
ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8, color='555555')

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
