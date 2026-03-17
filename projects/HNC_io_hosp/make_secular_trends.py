"""
make_secular_trends.py
Secular trend table: end-of-life care outcomes by year of death (2017–2023)
Output: C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\secular_trends.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\secular_trends.xlsx"

print("Loading io_analytic...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        death_year,
        hospice_enrolled,
        hospice_los_days,
        hospice_short_stay,
        io_within_14d_of_death,
        io_within_30d_of_death,
        in_hospital_death
    FROM io_analytic
""").df()
con.close()

for col in ['hospice_enrolled', 'hospice_short_stay', 'in_hospital_death',
            'io_within_14d_of_death', 'io_within_30d_of_death']:
    df[col] = df[col].fillna(0).astype(int)

N = len(df)
print(f"  Total N = {N:,}")

# ── Helper: compute stats for a group ─────────────────────────────────────────
def group_stats(g):
    n = len(g)
    hosp = g[g['hospice_enrolled'] == 1]
    n_hosp = len(hosp)

    pct = lambda num, denom: f"{num:,} ({100.0*num/denom:.1f}%)" if denom > 0 else "—"

    # Hospice enrollment
    hospice_pct = pct(n_hosp, n)

    # Hospice LOS: median (IQR) among enrolled
    los = hosp['hospice_los_days'].dropna()
    if len(los) > 0:
        los_med = f"{los.median():.0f} ({los.quantile(0.25):.0f}–{los.quantile(0.75):.0f})"
    else:
        los_med = "—"

    # Short stay ≤7 days (among enrolled)
    n_short = (hosp['hospice_short_stay'] == 1).sum()
    short_pct = pct(n_short, n_hosp)

    # IO ≤14d before death (all patients)
    n_14d = (g['io_within_14d_of_death'] == 1).sum()
    io14_pct = pct(n_14d, n)

    # IO ≤30d before death (all patients)
    n_30d = (g['io_within_30d_of_death'] == 1).sum()
    io30_pct = pct(n_30d, n)

    # In-hospital death (all patients)
    n_inh = (g['in_hospital_death'] == 1).sum()
    inh_pct = pct(n_inh, n)

    return {
        'N': f"{n:,}",
        'Hospice enrolled': hospice_pct,
        'Hospice LOS, median (IQR)': los_med,
        'Hospice LOS ≤7 days': short_pct,
        'IO ≤14d before death': io14_pct,
        'IO ≤30d before death': io30_pct,
        'In-hospital death': inh_pct,
    }

# ── Build table ────────────────────────────────────────────────────────────────
years = sorted(df['death_year'].dropna().unique())
rows = []

for yr in years:
    g = df[df['death_year'] == yr]
    stats = group_stats(g)
    row = {'Year': str(int(yr))}
    row.update(stats)
    rows.append(row)

# Total row
total_stats = group_stats(df)
total_row = {'Year': 'Total'}
total_row.update(total_stats)
rows.append(total_row)

df_out = pd.DataFrame(rows)
col_order = ['Year', 'N', 'Hospice enrolled', 'Hospice LOS, median (IQR)',
             'Hospice LOS ≤7 days', 'IO ≤14d before death',
             'IO ≤30d before death', 'In-hospital death']
df_out = df_out[col_order]

# ── Write Excel ────────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='BA0C2F')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='F5D0D6')
SECTION_FONT = Font(bold=True, size=10, color='7A0820')
ALT_FILL     = PatternFill('solid', fgColor='F9ECEE')
TOTAL_FILL   = PatternFill('solid', fgColor='F5D0D6')
TITLE_FONT   = Font(bold=True, size=13, color='BA0C2F')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Secular Trends'

ws.append([f'End-of-Life Care Outcomes by Year of Death (N = {N:,})'])
ws['A1'].font = TITLE_FONT
ws.append([])

header_row = ws.max_row + 1
for ci, cn in enumerate(col_order, 1):
    cell = ws.cell(row=header_row, column=ci, value=cn)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                wrap_text=True, vertical='center')
ws.row_dimensions[header_row].height = 48

for row_idx, row_series in enumerate(df_out.iterrows()):
    _, row_data = row_series
    ri = header_row + 1 + row_idx
    is_total = (row_data['Year'] == 'Total')
    for ci, col in enumerate(col_order, 1):
        val = row_data[col]
        cell = ws.cell(row=ri, column=ci, value=val)
        if is_total:
            cell.font = SECTION_FONT
            cell.fill = TOTAL_FILL
        elif row_idx % 2 == 1:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center', wrap_text=True)

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 10
for letter in ['C', 'D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[letter].width = 22

ws.freeze_panes = f'B{header_row + 1}'

footer_row = ws.max_row + 2
ws.cell(row=footer_row, column=1,
        value='LOS = length of stay. IO = immune checkpoint inhibitor. '
              'Hospice LOS ≤7 days reported among hospice enrollees. '
              'IO ≤14d and IO ≤30d before death reported in full cohort. '
              'Year = calendar year of death.')
ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8, color='555555')

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
