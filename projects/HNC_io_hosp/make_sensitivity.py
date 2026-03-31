"""
make_sensitivity.py
Sensitivity analysis: two parallel comparisons.

A) Regimen-adjusted IO discontinuation timing (io_analytic — primary cohort)
   Sheet 1 — Timing comparison: fixed vs regimen-adjusted, by estimated regimen
   Sheet 2 — Outcomes by regimen: hospice + in-hospital death rates
   Sheet 3 — Reclassified patients: those outside 30d window but within regimen window

B) Intent-to-treat comparison (io_analytic_itc)
   Sheet 4 — Primary vs ITC: key outcomes compared between cohorts
   Sheet 5 — ITC-only patients (de novo, no prior curative therapy): characteristics

Output: C:/Users/hsaee/Desktop/CMS_viewer/projects/HNC_io_hosp/sensitivity_analysis.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\sensitivity_analysis.xlsx"

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading io_analytic and io_analytic_itc...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        hospice_enrolled, in_hospital_death,
        days_last_io_to_death,
        io_within_14d_of_death, io_within_30d_of_death,
        median_interdose_days, estimated_regimen,
        discontinued_threshold_days, io_within_regimen_window,
        days_past_expected_dose_to_death
    FROM io_analytic
""").df()
df_itc = con.execute("""
    SELECT
        hospice_enrolled, in_hospital_death,
        days_last_io_to_death,
        io_within_14d_of_death, io_within_30d_of_death,
        io_within_regimen_window,
        days_past_expected_dose_to_death,
        had_prior_curative_therapy,
        hospice_los_days, hospice_short_stay,
        age_cat, sex, subsite_category, io_agent
    FROM io_analytic_itc
""").df()
con.close()

for col in ['hospice_enrolled','in_hospital_death','io_within_14d_of_death',
            'io_within_30d_of_death','io_within_regimen_window']:
    df[col] = df[col].fillna(0).astype(int)
df['estimated_regimen'] = df['estimated_regimen'].fillna('single-dose')

for col in ['hospice_enrolled','in_hospital_death','io_within_14d_of_death',
            'io_within_30d_of_death','io_within_regimen_window',
            'had_prior_curative_therapy','hospice_los_days','hospice_short_stay']:
    if col in df_itc.columns:
        df_itc[col] = df_itc[col].fillna(0)
for col in ['hospice_enrolled','in_hospital_death','io_within_14d_of_death',
            'io_within_30d_of_death','io_within_regimen_window',
            'had_prior_curative_therapy','hospice_short_stay']:
    df_itc[col] = df_itc[col].astype(int)

N    = len(df)
N_itc = len(df_itc)
df_primary  = df_itc[df_itc['had_prior_curative_therapy'] == 1]
df_itc_only = df_itc[df_itc['had_prior_curative_therapy'] == 0]
print(f"  Primary cohort N = {N:,}")
print(f"  ITC cohort N     = {N_itc:,}  (ITC-only = {len(df_itc_only):,})")
print(f"  Regimen distribution:\n{df['estimated_regimen'].value_counts()}")

# ── Helpers ────────────────────────────────────────────────────────────────────
REGIMEN_ORDER = ['q2w', 'q3w', 'q6w', 'single-dose', 'other']
REGIMEN_LABELS = {
    'q2w':         'Every 2 weeks (q2w)',
    'q3w':         'Every 3 weeks (q3w)',
    'q6w':         'Every 6 weeks (q6w)',
    'single-dose': 'Single dose only',
    'other':       'Other / irregular',
}

def med_iqr(series):
    s = series.dropna()
    if len(s) == 0:
        return '—'
    return f"{s.median():.0f} ({s.quantile(0.25):.0f}–{s.quantile(0.75):.0f})"

def n_pct(n, total):
    if total == 0:
        return '—'
    return f"{n:,} ({100.0 * n / total:.1f}%)"

# ── Sheet 1: Timing comparison ─────────────────────────────────────────────────
print("Building Sheet 1: Timing comparison...")

s1_rows = []

def s1_row(label, sub):
    n         = len(sub)
    mdn       = med_iqr(sub['days_last_io_to_death'])
    w14       = n_pct(sub['io_within_14d_of_death'].sum(), n)
    w30       = n_pct(sub['io_within_30d_of_death'].sum(), n)
    wreg      = n_pct(sub['io_within_regimen_window'].sum(), n)
    med_idose = med_iqr(sub['median_interdose_days'])
    mdn_past  = med_iqr(sub['days_past_expected_dose_to_death'])
    s1_rows.append({
        'Estimated Dosing Regimen':                                    label,
        'N':                                                           n,
        'Median time between IO doses, days (IQR)':                    med_idose,
        'Days from last IO dose to death, median (IQR)':               mdn,
        'Days from next expected dose to death, median (IQR)†':        mdn_past,
        'Last IO dose within 14 days of death, n (%)':                 w14,
        'Last IO dose within 30 days of death, n (%)':                 w30,
        'Last IO dose within regimen-adjusted window‡, n (%)':         wreg,
    })

s1_row('Overall', df)
for reg in REGIMEN_ORDER:
    sub = df[df['estimated_regimen'] == reg]
    if len(sub) > 0:
        s1_row(f'  {REGIMEN_LABELS[reg]}', sub)

df_s1 = pd.DataFrame(s1_rows)

# ── Sheet 2: Outcomes by regimen ───────────────────────────────────────────────
print("Building Sheet 2: Outcomes by regimen...")

s2_rows = []

def s2_row(label, sub):
    n    = len(sub)
    hosp = n_pct(sub['hospice_enrolled'].sum(), n)
    ihd  = n_pct(sub['in_hospital_death'].sum(), n)
    s2_rows.append({
        'Estimated Dosing Regimen':  label,
        'N':                         n,
        'Enrolled in hospice, n (%)': hosp,
        'Died in hospital, n (%)':   ihd,
    })

s2_row('Overall', df)
for reg in REGIMEN_ORDER:
    sub = df[df['estimated_regimen'] == reg]
    if len(sub) > 0:
        s2_row(f'  {REGIMEN_LABELS[reg]}', sub)

df_s2 = pd.DataFrame(s2_rows)

# ── Sheet 3: Reclassified patients ─────────────────────────────────────────────
print("Building Sheet 3: Reclassified patients...")

# Patients outside 30d window but inside their regimen window
reclassified = df[(df['io_within_30d_of_death'] == 0) & (df['io_within_regimen_window'] == 1)]
Nr = len(reclassified)
print(f"  Reclassified N = {Nr:,} ({100.0*Nr/N:.1f}% of cohort)")

COL_GROUP = 'Patient Group (by IO timing definition)'
COL_HOSP  = 'Enrolled in hospice, n (%)'
COL_IHD   = 'Died in hospital, n (%)'
COL_DAYS  = 'Days from last IO dose to death, median (IQR)'
COL_PAST  = 'Days from next expected dose to death, median (IQR)†'
COL_NOTE  = 'Interpretation'

s3_rows = []
n_outside_30d  = (df['io_within_30d_of_death'] == 0).sum()
n_confirmed_dc = n_outside_30d - Nr
df_gt30        = df[df['io_within_30d_of_death'] == 0]
df_confirmed   = df[(df['io_within_30d_of_death'] == 0) & (df['io_within_regimen_window'] == 0)]

s3_rows.append({
    COL_GROUP: 'Last IO dose >30 days before death  [conventional definition: IO discontinued]',
    'N':       n_outside_30d,
    COL_HOSP:  n_pct(df_gt30['hospice_enrolled'].sum(), n_outside_30d),
    COL_IHD:   n_pct(df_gt30['in_hospital_death'].sum(), n_outside_30d),
    COL_DAYS:  med_iqr(df_gt30['days_last_io_to_death']),
    COL_PAST:  med_iqr(df_gt30['days_past_expected_dose_to_death']),
    COL_NOTE:  'Primary analysis group',
})
s3_rows.append({
    COL_GROUP: '  Of whom: died before next expected dose  [regimen-adjusted: still on IO at death]',
    'N':       Nr,
    COL_HOSP:  n_pct(reclassified['hospice_enrolled'].sum(), Nr),
    COL_IHD:   n_pct(reclassified['in_hospital_death'].sum(), Nr),
    COL_DAYS:  med_iqr(reclassified['days_last_io_to_death']),
    COL_PAST:  med_iqr(reclassified['days_past_expected_dose_to_death']),
    COL_NOTE:  'Reclassified as on-treatment by this sensitivity analysis',
})
s3_rows.append({
    COL_GROUP: '  Of whom: died after next expected dose  [confirmed IO discontinuation]',
    'N':       n_confirmed_dc,
    COL_HOSP:  n_pct(df_confirmed['hospice_enrolled'].sum(), n_confirmed_dc),
    COL_IHD:   n_pct(df_confirmed['in_hospital_death'].sum(), n_confirmed_dc),
    COL_DAYS:  med_iqr(df_confirmed['days_last_io_to_death']),
    COL_PAST:  med_iqr(df_confirmed['days_past_expected_dose_to_death']),
    COL_NOTE:  'Truly discontinued IO before death',
})
for reg in REGIMEN_ORDER:
    sub = reclassified[reclassified['estimated_regimen'] == reg]
    if len(sub) > 0:
        s3_rows.append({
            COL_GROUP: f'    Reclassified as still on IO — {REGIMEN_LABELS[reg]}',
            'N':        len(sub),
            COL_HOSP:  n_pct(sub['hospice_enrolled'].sum(), len(sub)),
            COL_IHD:   n_pct(sub['in_hospital_death'].sum(), len(sub)),
            COL_DAYS:  med_iqr(sub['days_last_io_to_death']),
            COL_PAST:  med_iqr(sub['days_past_expected_dose_to_death']),
            COL_NOTE:  '',
        })

df_s3 = pd.DataFrame(s3_rows)

# ── Sheet 4: Primary vs ITC comparison ────────────────────────────────────────
print("Building Sheet 4: Primary vs ITC comparison...")

def compare_row(label, sub_p, sub_i):
    np_, ni_ = len(sub_p), len(sub_i)
    return {
        'Outcome':                              label,
        f'Primary cohort (N={N:,})':            n_pct(sub_p['hospice_enrolled'].sum(), np_) if 'hospice' in label.lower()
                                                 else n_pct(sub_p['in_hospital_death'].sum(), np_) if 'in-hospital' in label.lower()
                                                 else n_pct(sub_p['io_within_14d_of_death'].sum(), np_) if '14d' in label.lower()
                                                 else n_pct(sub_p['io_within_30d_of_death'].sum(), np_) if '30d' in label.lower()
                                                 else n_pct(sub_p['io_within_regimen_window'].sum(), np_) if 'regimen' in label.lower()
                                                 else med_iqr(sub_p['days_last_io_to_death']),
        f'ITC cohort (N={N_itc:,})':            n_pct(sub_i['hospice_enrolled'].sum(), ni_) if 'hospice' in label.lower()
                                                 else n_pct(sub_i['in_hospital_death'].sum(), ni_) if 'in-hospital' in label.lower()
                                                 else n_pct(sub_i['io_within_14d_of_death'].sum(), ni_) if '14d' in label.lower()
                                                 else n_pct(sub_i['io_within_30d_of_death'].sum(), ni_) if '30d' in label.lower()
                                                 else n_pct(sub_i['io_within_regimen_window'].sum(), ni_) if 'regimen' in label.lower()
                                                 else med_iqr(sub_i['days_last_io_to_death']),
    }

P_COL = f'Prior ITC\n(prior curative therapy, N={N:,})'
I_COL = f'ITC + No ITC\n(all HNC+IO patients, N={N_itc:,})'

def row4(outcome, p_val, i_val):
    return {'Outcome / Characteristic': outcome, P_COL: p_val, I_COL: i_val}

s4_rows = [
    row4('Total patients, N', f'{N:,}', f'{N_itc:,}'),
    row4('— OUTCOMES —', '', ''),
    row4('Enrolled in hospice, n (%)',
         n_pct(df['hospice_enrolled'].sum(), N),
         n_pct(df_itc['hospice_enrolled'].sum(), N_itc)),
    row4('Hospice length of stay among enrolled, median days (IQR)',
         med_iqr(df_primary['hospice_los_days'][df_primary['hospice_enrolled']==1]) if 'hospice_los_days' in df_primary.columns else '—',
         med_iqr(df_itc['hospice_los_days'][df_itc['hospice_enrolled']==1]) if 'hospice_los_days' in df_itc.columns else '—'),
    row4('Died in hospital, n (%)',
         n_pct(df['in_hospital_death'].sum(), N),
         n_pct(df_itc['in_hospital_death'].sum(), N_itc)),
    row4('— IO TIMING —', '', ''),
    row4('Days from last IO dose to death, median (IQR)',
         med_iqr(df['days_last_io_to_death']),
         med_iqr(df_itc['days_last_io_to_death'])),
    row4('Last IO dose within 14 days of death, n (%)',
         n_pct(df['io_within_14d_of_death'].sum(), N),
         n_pct(df_itc['io_within_14d_of_death'].sum(), N_itc)),
    row4('Last IO dose within 30 days of death, n (%)',
         n_pct(df['io_within_30d_of_death'].sum(), N),
         n_pct(df_itc['io_within_30d_of_death'].sum(), N_itc)),
    row4('Last IO dose within regimen-adjusted window†, n (%)\n(likely still on IO at death)',
         n_pct(df['io_within_regimen_window'].sum(), N),
         n_pct(df_itc['io_within_regimen_window'].sum(), N_itc)),
]
df_s4 = pd.DataFrame(s4_rows)

# ── Sheet 5: Prior ITC vs No ITC comparison ───────────────────────────────────
print("Building Sheet 5: Prior ITC vs No ITC comparison...")

Nd  = len(df_itc_only)
Np  = len(df_primary)
P5_COL = f'Prior ITC\n(prior curative therapy, N={Np:,})'
N5_COL = f'No ITC\n(no prior curative therapy, N={Nd:,})'
s5_rows = []

def s5_add(label, p_val, n_val):
    s5_rows.append({'Characteristic': label, P5_COL: p_val, N5_COL: n_val})

def s5_npct2(p_series, n_series, val=1):
    return (n_pct((p_series == val).sum(), len(p_series)),
            n_pct((n_series == val).sum(), len(n_series)))

s5_add('Total patients, N', f'{Np:,}', f'{Nd:,}')
s5_add('OUTCOMES', '', '')
pv, nv = s5_npct2(df_primary['hospice_enrolled'], df_itc_only['hospice_enrolled'])
s5_add('  Enrolled in hospice, n (%)', pv, nv)
pv, nv = s5_npct2(df_primary['in_hospital_death'], df_itc_only['in_hospital_death'])
s5_add('  Died in hospital, n (%)', pv, nv)
s5_add('  Days from last IO dose to death, median (IQR)',
       med_iqr(df_primary['days_last_io_to_death']),
       med_iqr(df_itc_only['days_last_io_to_death']))
pv, nv = s5_npct2(df_primary['io_within_14d_of_death'], df_itc_only['io_within_14d_of_death'])
s5_add('  Last IO dose within 14 days of death, n (%)', pv, nv)
pv, nv = s5_npct2(df_primary['io_within_30d_of_death'], df_itc_only['io_within_30d_of_death'])
s5_add('  Last IO dose within 30 days of death, n (%)', pv, nv)
pv, nv = s5_npct2(df_primary['io_within_regimen_window'], df_itc_only['io_within_regimen_window'])
s5_add('  Last IO dose within regimen-adjusted window‡, n (%)', pv, nv)
s5_add('  Days from next expected dose to death, median (IQR)†',
       med_iqr(df_primary['days_past_expected_dose_to_death']),
       med_iqr(df_itc_only['days_past_expected_dose_to_death']))
s5_add('DEMOGRAPHICS', '', '')
for cat in ['66-69','70-74','75-79','80-84','85+']:
    s5_add(f'  Age {cat}, n (%)',
           n_pct((df_primary['age_cat']==cat).sum(), Np),
           n_pct((df_itc_only['age_cat']==cat).sum(), Nd))
s5_add('  Male sex, n (%)',
       n_pct((df_primary['sex']=='Male').sum(), Np),
       n_pct((df_itc_only['sex']=='Male').sum(), Nd))
s5_add('TUMOR SUBSITE', '', '')
all_subsites = sorted(set(df_primary['subsite_category'].dropna()) |
                      set(df_itc_only['subsite_category'].dropna()))
for sub in all_subsites:
    s5_add(f'  {sub}, n (%)',
           n_pct((df_primary['subsite_category']==sub).sum(), Np),
           n_pct((df_itc_only['subsite_category']==sub).sum(), Nd))
s5_add('IO AGENT', '', '')
all_agents = sorted(set(df_primary['io_agent'].dropna()) |
                    set(df_itc_only['io_agent'].dropna()))
for ag in all_agents:
    s5_add(f'  {ag}, n (%)',
           n_pct((df_primary['io_agent']==ag).sum(), Np),
           n_pct((df_itc_only['io_agent']==ag).sum(), Nd))

df_s5 = pd.DataFrame(s5_rows)

# ── Excel styling ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='BA0C2F')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='F5D0D6')
SECTION_FONT = Font(bold=True, size=10, color='7A0820')
ALT_FILL     = PatternFill('solid', fgColor='F9ECEE')
TITLE_FONT   = Font(bold=True, size=13, color='BA0C2F')
NOTE_FONT    = Font(italic=True, size=8, color='555555')


def write_sheet(wb, sheet_name, title, df_data, col_widths):
    ws = wb.create_sheet(title=sheet_name)
    ws.append([title])
    ws['A1'].font = TITLE_FONT
    ws.append([])

    header_row = ws.max_row + 1
    for ci, col in enumerate(df_data.columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                   wrap_text=True, vertical='center')

    alt = 0
    for ri, row_vals in enumerate(df_data.itertuples(index=False), header_row + 1):
        first_val = str(row_vals[0])
        is_overall  = first_val.strip().lower() in ('overall', 'total patients, n')
        is_section  = first_val.strip().isupper() and not first_val.startswith(' ')
        is_dash_sep = first_val.strip().startswith('—')
        is_indent   = first_val.startswith('  ')
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if is_section or is_dash_sep:
                cell.font = SECTION_FONT
                cell.fill = SECTION_FILL
            elif is_overall:
                cell.font = Font(bold=True, size=10)
            elif is_indent:
                alt += 1
                if alt % 2 == 0:
                    cell.fill = ALT_FILL
            cell.alignment = Alignment(
                horizontal='left' if ci == 1 else 'center',
                vertical='center', wrap_text=(ci == 1))

    for ci, width in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)].width = width
    ws.freeze_panes = f'B{header_row + 1}'
    return ws


print(f"Writing {OUT_PATH} ...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

FOOTNOTE_REGIMEN = (
    '† Days from next expected dose to death = days from last IO dose to death minus median interdose interval. '
    'Negative values indicate the patient died before their next dose was due (still on schedule); '
    'positive values indicate the patient had already missed at least one dose. '
    'Single-dose patients use a default interval of 42 days. '
    '‡ Regimen-adjusted window = median interdose interval + 14 days tolerance. '
    'A patient within this window likely died before their next scheduled infusion.'
)
FOOTNOTE_ITC = (
    'Prior ITC (N=2,527): patients who received prior curative-intent therapy (surgery or radiation) '
    'before IO — i.e., recurrent or metastatic HNC. '
    'No ITC (N=817): patients with NO documented prior curative therapy — likely de novo metastatic '
    'or unresectable HNC treated with IO as first-line. '
    'ITC + No ITC (N=3,344): all HNC+IO patients combined; includes every patient in the Prior ITC '
    'cohort plus the No ITC group.'
)

ws1 = write_sheet(
    wb, 'A1. IO Timing by Regimen',
    f'Sensitivity A: Does the dosing regimen affect how we classify IO timing near death? (Primary cohort, N = {N:,})',
    df_s1,
    [38, 7, 38, 36, 34, 34, 40]
)
ws1.cell(row=ws1.max_row + 2, column=1, value=FOOTNOTE_REGIMEN).font = NOTE_FONT

write_sheet(
    wb, 'A2. Outcomes by Regimen',
    f'Sensitivity A: Do hospice and in-hospital death rates differ by dosing regimen? (Primary cohort, N = {N:,})',
    df_s2,
    [38, 7, 28, 28]
)

ws3 = write_sheet(
    wb, 'A3. Reclassified Patients',
    (f'Sensitivity A: Among patients classified as "IO discontinued" by the 30-day rule, '
     f'how many were actually still within their dosing cycle? (Primary cohort, N = {N:,})'),
    df_s3,
    [68, 7, 28, 28, 38, 44]
)
ws3.cell(row=ws3.max_row + 2, column=1, value=FOOTNOTE_REGIMEN).font = NOTE_FONT

ws4 = write_sheet(
    wb, 'B1. Prior ITC vs ITC+No ITC',
    (f'Sensitivity B: Prior ITC only (N={N:,}) vs ITC + No ITC combined (N={N_itc:,}) — '
     f'does including No ITC patients change the results?'),
    df_s4,
    [52, 36, 36]
)
ws4.cell(row=ws4.max_row + 2, column=1, value=FOOTNOTE_ITC).font = NOTE_FONT
ws4.cell(row=ws4.max_row + 1, column=1, value=FOOTNOTE_REGIMEN).font = NOTE_FONT

ws5 = write_sheet(
    wb, 'B2. Prior ITC vs No ITC',
    (f'Sensitivity B: Direct comparison of Prior ITC (N={Np:,}) vs No ITC (N={Nd:,}) patients — '
     f'does the absence of prior curative therapy change end-of-life care patterns?'),
    df_s5,
    [52, 36, 36]
)
ws5.cell(row=ws5.max_row + 2, column=1, value=FOOTNOTE_ITC).font = NOTE_FONT

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
