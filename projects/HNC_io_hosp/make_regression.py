"""
make_regression.py
Table 3: Multivariable logistic regression — hospice enrollment (primary outcome)
Table 4: Multivariable logistic regression — in-hospital death (secondary outcome)
Output: C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\table3_regression.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
import statsmodels.formula.api as smf
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\table3_regression.xlsx"

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading io_analytic...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        hospice_enrolled, in_hospital_death,
        age_at_death, sex, race,
        dual_eligible, urban_rural, census_region,
        subsite_category, io_agent, io_regimen,
        last_episode_doses,
        van_walraven_score,
        primary_curative_type,
        death_year
    FROM io_analytic
""").df()
con.close()

# ── Data prep ──────────────────────────────────────────────────────────────────
df['hospice_enrolled']  = df['hospice_enrolled'].fillna(0).astype(int)
df['in_hospital_death'] = df['in_hospital_death'].fillna(0).astype(int)
df['dual_eligible']     = df['dual_eligible'].fillna(0).astype(int)
df['van_walraven_score'] = pd.to_numeric(df['van_walraven_score'], errors='coerce').fillna(0)

# Fill unknown geography (non-US addresses, CT restructuring) rather than dropping
df['urban_rural']   = df['urban_rural'].fillna('Unknown')
df['census_region'] = df['census_region'].fillna('Unknown')

# Collapse sparse race categories
df['race_collapsed'] = df['race'].replace({
    'Asian/PI': 'Other/Unknown',
    'Native American': 'Other/Unknown',
    'Other': 'Other/Unknown',
    'Unknown': 'Other/Unknown',
})

# Set reference categories using categorical dtype
cat_vars = {
    'sex':                ('Male', ['Male', 'Female']),
    'race_collapsed':     ('White', ['White', 'Black', 'Hispanic', 'Other/Unknown']),
    'urban_rural':        ('Metro', ['Metro', 'Non-metro', 'Unknown']),
    'census_region':      ('South', ['Northeast', 'Midwest', 'South', 'West', 'Unknown']),
    'subsite_category':   None,
    'io_agent':           ('pembrolizumab', None),
    'io_regimen':         ('IO monotherapy', ['IO monotherapy', 'chemo-IO']),
    'primary_curative_type': ('radiation', None),
}

for col, spec in cat_vars.items():
    if spec is None:
        df[col] = pd.Categorical(df[col])
    else:
        ref, order = spec
        if order is None:
            df[col] = pd.Categorical(df[col])
        else:
            # Only keep levels that exist
            exist = [o for o in order if o in df[col].unique()]
            df[col] = pd.Categorical(df[col], categories=exist)
        # Reorder so ref is first
        if spec is not None and spec[0] in df[col].cat.categories:
            cats = [spec[0]] + [c for c in df[col].cat.categories if c != spec[0]]
            df[col] = df[col].cat.reorder_categories(cats)

df['age_at_death'] = pd.to_numeric(df['age_at_death'], errors='coerce')
df['last_episode_doses'] = pd.to_numeric(df['last_episode_doses'], errors='coerce')
df['death_year_c'] = df['death_year'].astype(float) - 2017  # centered at 2017

# Drop rows with any missing predictor
predictors = ['age_at_death', 'sex', 'race_collapsed', 'dual_eligible', 'urban_rural',
              'census_region', 'subsite_category', 'io_agent', 'io_regimen',
              'last_episode_doses', 'van_walraven_score',
              'primary_curative_type', 'death_year_c']

df_model = df[predictors + ['hospice_enrolled', 'in_hospital_death']].dropna()
print(f"  Model dataset: {len(df_model):,} rows (dropped {len(df)-len(df_model):,} with missing predictors)")

# ── Fit models ────────────────────────────────────────────────────────────────
formula = (
    'hospice_enrolled ~ age_at_death + C(sex) + C(race_collapsed) + dual_eligible '
    '+ C(urban_rural) + C(census_region) + C(subsite_category) '
    '+ C(io_agent) + C(io_regimen) '
    '+ last_episode_doses + van_walraven_score + C(primary_curative_type) '
    '+ death_year_c'
)

print("Fitting Table 3 (hospice enrollment)...")
model3 = smf.logit(formula.replace('hospice_enrolled', 'hospice_enrolled'), data=df_model).fit(
    method='newton', maxiter=500, disp=False)

print("Fitting Table 4 (in-hospital death)...")
model4 = smf.logit(formula.replace('hospice_enrolled', 'in_hospital_death'), data=df_model).fit(
    method='newton', maxiter=500, disp=False)

# ── Extract results ────────────────────────────────────────────────────────────
def extract_results(model):
    """Return DataFrame with term, OR, 95% CI, p-value."""
    coef = model.params
    conf = model.conf_int()
    pval = model.pvalues
    results = pd.DataFrame({
        'term': coef.index,
        'coef': coef.values,
        'ci_lo': conf[0].values,
        'ci_hi': conf[1].values,
        'pval': pval.values,
    })
    results['OR']    = np.exp(results['coef'])
    results['CI_lo'] = np.exp(results['ci_lo'])
    results['CI_hi'] = np.exp(results['ci_hi'])
    results = results[results['term'] != 'Intercept']
    return results

res3 = extract_results(model3)
res4 = extract_results(model4)

# ── Format results ────────────────────────────────────────────────────────────
def fmt_or(row):
    return f"{row['OR']:.2f} ({row['CI_lo']:.2f}–{row['CI_hi']:.2f})"

def fmt_p(p):
    if p < 0.001:
        return '<0.001'
    return f'{p:.3f}'

def clean_term(t):
    """Convert statsmodels term name to readable label."""
    t = t.replace('C(sex)[T.', 'Sex: ').replace('C(race)[T.', 'Race: ')
    t = t.replace('C(urban_rural)[T.', 'Urban/rural: ').replace('C(census_region)[T.', 'Region: ')
    t = t.replace('C(subsite_category)[T.', 'Subsite: ').replace('C(io_agent)[T.', 'IO agent: ')
    t = t.replace('C(io_regimen)[T.', 'IO regimen: ')
    t = t.replace('C(elixhauser_cat)[T.', 'Elixhauser: ').replace('C(primary_curative_type)[T.', 'Prior tx: ')
    t = t.replace(']', '')
    t = t.replace('age_at_death', 'Age at death (per year)')
    t = t.replace('dual_eligible', 'Dual eligible')
    t = t.replace('last_episode_doses', 'Last episode IO doses (per dose)')
    t = t.replace('death_year_c', 'Calendar year (per year from 2017)')
    return t

for res in [res3, res4]:
    res['label']  = res['term'].apply(clean_term)
    res['or_ci']  = res.apply(fmt_or, axis=1)
    res['p_fmt']  = res['pval'].apply(fmt_p)

# ── Write Excel ────────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL = PatternFill('solid', fgColor='BA0C2F')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL    = PatternFill('solid', fgColor='F9ECEE')
SIG_FILL    = PatternFill('solid', fgColor='FFE699')
TITLE_FONT  = Font(bold=True, size=13, color='BA0C2F')
REF_FONT    = Font(italic=True, size=9, color='888888')

wb = openpyxl.Workbook()

for sheet_name, res, outcome_label, n_outcome in [
    ('Table 3 - Hospice',  res3, 'Hospice Enrollment', df_model['hospice_enrolled'].sum()),
    ('Table 4 - In-Hosp Death', res4, 'In-Hospital Death', df_model['in_hospital_death'].sum()),
]:
    ws = wb.create_sheet(title=sheet_name)

    ws.append([f'Multivariable Logistic Regression: {outcome_label} '
               f'(n events = {int(n_outcome):,} / {len(df_model):,})'])
    ws['A1'].font = TITLE_FONT
    ws.append([])

    # Split into significant and non-significant
    res_sig   = res[res['pval'] < 0.05].copy()
    res_ns    = res[res['pval'] >= 0.05].copy()
    ns_labels = ', '.join(res_ns['label'].tolist())

    hr = ws.max_row + 1
    for ci, cn in enumerate(['Variable', 'OR (95% CI)', 'p-value'], 1):
        cell = ws.cell(row=hr, column=ci, value=cn)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                    wrap_text=True, vertical='center')

    for i, row in enumerate(res_sig.itertuples(index=False), hr + 1):
        for ci, val in enumerate([row.label, row.or_ci, row.p_fmt], 1):
            cell = ws.cell(row=i, column=ci, value=val)
            cell.alignment = Alignment(
                horizontal='left' if ci == 1 else 'center', vertical='center')
            cell.fill = SIG_FILL

    ws.column_dimensions['A'].width = 44
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 12
    ws.freeze_panes = f'B{hr + 1}'

    footer = ws.max_row + 2
    ws.cell(row=footer, column=1,
            value=f'Showing only variables with p<0.05 (n={len(res_sig)} of {len(res)} terms). '
                  f'Model also adjusted for: {ns_labels}. '
                  f'OR = odds ratio; CI = 95% confidence interval. '
                  f'Reference: Male, White, Metro, South, IO monotherapy, radiation (curative).')
    ws.cell(row=footer, column=1).font = Font(italic=True, size=8, color='555555')

# Remove default empty sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"\nModel 3 (hospice): pseudo-R² = {model3.prsquared:.3f}, AIC = {model3.aic:.1f}")
print(f"Model 4 (in-hosp): pseudo-R² = {model4.prsquared:.3f}, AIC = {model4.aic:.1f}")
