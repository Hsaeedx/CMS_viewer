"""
make_table2.py
Hospice utilization and end-of-life care patterns
Output: C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\table2.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
import scipy.stats as stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\table2.xlsx"

print("Loading io_analytic...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        hospice_enrolled, hospice_los_days, hospice_short_stay,
        days_last_io_to_hospice, days_last_io_to_death,
        days_last_io_to_death_cat,
        in_hospital_death,
        death_year,
        io_agent, io_regimen,
        subsite_category, age_cat
    FROM io_analytic
""").df()
con.close()

df['hospice_enrolled']  = df['hospice_enrolled'].fillna(0).astype(int)
df['in_hospital_death'] = df['in_hospital_death'].fillna(0).astype(int)
df['hospice_short_stay']= df['hospice_short_stay'].fillna(0).astype(int)
hosp = df[df['hospice_enrolled'] == 1]
N    = len(df)
Nh   = len(hosp)

print(f"  Total={N:,}  Hospice={Nh:,} ({100*Nh/N:.1f}%)")

# ── Build rows ────────────────────────────────────────────────────────────────
rows = []

def section(title):
    rows.append({'label': title, 'is_section': True, 'val': '', 'sub': ''})

def add(label, val, sub='', indent=False):
    rows.append({
        'label': ('    ' + label) if indent else label,
        'is_section': False,
        'val': val,
        'sub': sub,
    })

def n_pct(series, val=None):
    if val is not None:
        n = (series == val).sum()
    else:
        n = series.sum()
    p = 100.0 * n / len(series)
    return f"{n:,} ({p:.1f}%)", n

def med_iqr(series):
    s = series.dropna()
    return f"{s.median():.0f} ({s.quantile(0.25):.0f}–{s.quantile(0.75):.0f})"

# ── Primary outcome ───────────────────────────────────────────────────────────
section('PRIMARY OUTCOME')
v, _ = n_pct(df['hospice_enrolled'], 1)
add('Hospice enrolled', v)

# ── Hospice utilization (among enrolled) ─────────────────────────────────────
section('HOSPICE UTILIZATION (among enrolled, n = {:,})'.format(Nh))
add('Hospice LOS, median (IQR)', med_iqr(hosp['hospice_los_days']), 'days')
v, _ = n_pct(hosp['hospice_short_stay'], 1)
add('Short stay ≤7 days', v)
add('Days last IO → hospice, median (IQR)',
    med_iqr(hosp['days_last_io_to_hospice']), 'days')

# ── Timing last IO to death ───────────────────────────────────────────────────
section('TIMING: LAST IO DOSE TO DEATH')
add('Days last IO → death, median (IQR)',
    med_iqr(df['days_last_io_to_death']), 'days')
for cat, label in [
    ('<=3 days',  '≤3 days'),
    ('4-14 days', '4–14 days'),
    ('15-30 days','15–30 days'),
    ('31-90 days','31–90 days'),
    ('>90 days',  '>90 days'),
]:
    n = (df['days_last_io_to_death_cat'] == cat).sum()
    p = 100.0 * n / N
    add(label, f"{n:,} ({p:.1f}%)", indent=True)

# ── Secondary outcomes ────────────────────────────────────────────────────────
section('SECONDARY OUTCOMES')
v, _ = n_pct(df['in_hospital_death'], 1)
add('In-hospital death', v)

# ── Write Excel ───────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='BA0C2F')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='F5D0D6')
SECTION_FONT = Font(bold=True, size=10, color='7A0820')
ALT_FILL     = PatternFill('solid', fgColor='F9ECEE')
TITLE_FONT   = Font(bold=True, size=13, color='BA0C2F')

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Table 2'

ws.append([f'Table 2. Hospice Utilization and End-of-Life Care Patterns (N = {N:,})'])
ws['A1'].font = TITLE_FONT
ws.append([])

header_row = ws.max_row + 1
for ci, cn in enumerate(['Characteristic', f'N = {N:,}', 'Notes'], 1):
    cell = ws.cell(row=header_row, column=ci, value=cn)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='left' if ci == 1 else 'center',
                                wrap_text=True, vertical='center')

alt = 0
for ri, row_data in enumerate(pd.DataFrame(rows).itertuples(index=False), header_row + 1):
    is_sec = row_data.is_section
    if not is_sec:
        alt += 1
    for ci, val in enumerate([row_data.label, row_data.val, row_data.sub], 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if is_sec:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif alt % 2 == 0:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center', wrap_text=(ci == 1))

ws.column_dimensions['A'].width = 52
ws.column_dimensions['B'].width = 24
ws.column_dimensions['C'].width = 12
ws.freeze_panes = f'B{header_row + 1}'

footer_row = ws.max_row + 2
ws.cell(row=footer_row, column=1,
        value='LOS = length of stay. IO = immune checkpoint inhibitor. '
              'Short hospice stay defined as LOS ≤7 days.')
ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8, color='555555')

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
