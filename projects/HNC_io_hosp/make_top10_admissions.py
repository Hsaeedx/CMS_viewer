"""
make_top10_admissions.py
Top 10 principal diagnoses for inpatient admissions within 30 days of death
among IO hospice cohort (io_analytic, N=2,527).

Sheet 1: Raw top 10 ICD-10 codes (full code)
Sheet 2: Top 10 consolidated by 3-character ICD-10 category

Output: top10_admissions.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"C:\Users\hsaee\Desktop\CMS_viewer\projects\HNC_io_hosp\top10_admissions.xlsx"

# ── ICD-10 descriptions ────────────────────────────────────────────────────────
DESCRIPTIONS_FULL = {
    "A419":  "Sepsis, unspecified organism",
    "J690":  "Aspiration pneumonitis",
    "J9601": "Acute respiratory failure, unspecified",
    "J189":  "Pneumonia, unspecified organism",
    "C329":  "Malignant neoplasm of larynx, unspecified",
    "J9621": "Acute and chronic respiratory failure with hypoxia",
    "E860":  "Dehydration",
    "U071":  "COVID-19",
    "N179":  "Acute kidney failure, unspecified",
    "G893":  "Neoplasm-related pain",
}

DESCRIPTIONS_3 = {
    "A41": "Sepsis",
    "J69": "Pneumonitis due to solids and liquids (aspiration)",
    "J96": "Respiratory failure",
    "J18": "Pneumonia, unspecified organism",
    "C32": "Malignant neoplasm of larynx",
    "E86": "Volume depletion (dehydration)",
    "U07": "COVID-19",
    "N17": "Acute kidney failure",
    "G89": "Pain, not elsewhere classified",
    "C34": "Malignant neoplasm of bronchus and lung",
    "J44": "Chronic obstructive pulmonary disease",
    "R65": "Systemic inflammatory response / sepsis",
    "I50": "Heart failure",
    "K92": "Other diseases of digestive system",
    "C10": "Malignant neoplasm of oropharynx",
    "C01": "Malignant neoplasm of base of tongue",
    "C78": "Secondary malignant neoplasm of respiratory/digestive organs",
    "C79": "Secondary malignant neoplasm of other sites",
    "J95": "Postprocedural respiratory complications",
    "C02": "Malignant neoplasm of tongue (other/unspecified)",
}

# ── Colors ─────────────────────────────────────────────────────────────────────
SCARLET     = "BA0C2F"
WHITE       = "FFFFFF"
SCARLET_D40 = "70071C"
LIGHT_GRAY  = "F2F2F2"

THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    left=Side(style="thin", color="DDDDDD"),
)


# ── Query ──────────────────────────────────────────────────────────────────────
print("Querying database...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")

sql_full = """
WITH total AS (
    SELECT COUNT(*) AS total_n
    FROM io_inp_claims i
    JOIN io_analytic a ON i.DSYSRTKY = a.DSYSRTKY
    WHERE TRY_STRPTIME(i.ADMSN_DT, '%Y%m%d') BETWEEN (a.death_dt - INTERVAL 30 DAY) AND a.death_dt
      AND i.PRNCPAL_DGNS_CD IS NOT NULL AND i.PRNCPAL_DGNS_CD <> ''
)
SELECT i.PRNCPAL_DGNS_CD AS icd_code,
       COUNT(*) AS n,
       ROUND(100.0 * COUNT(*) / MAX(t.total_n), 1) AS pct
FROM io_inp_claims i
JOIN io_analytic a ON i.DSYSRTKY = a.DSYSRTKY
CROSS JOIN total t
WHERE TRY_STRPTIME(i.ADMSN_DT, '%Y%m%d') BETWEEN (a.death_dt - INTERVAL 30 DAY) AND a.death_dt
  AND i.PRNCPAL_DGNS_CD IS NOT NULL AND i.PRNCPAL_DGNS_CD <> ''
GROUP BY i.PRNCPAL_DGNS_CD
ORDER BY n DESC;
"""

sql_3char = """
WITH total AS (
    SELECT COUNT(*) AS total_n
    FROM io_inp_claims i
    JOIN io_analytic a ON i.DSYSRTKY = a.DSYSRTKY
    WHERE TRY_STRPTIME(i.ADMSN_DT, '%Y%m%d') BETWEEN (a.death_dt - INTERVAL 30 DAY) AND a.death_dt
      AND i.PRNCPAL_DGNS_CD IS NOT NULL AND i.PRNCPAL_DGNS_CD <> ''
)
SELECT LEFT(i.PRNCPAL_DGNS_CD, 3) AS icd3,
       COUNT(*) AS n,
       ROUND(100.0 * COUNT(*) / MAX(t.total_n), 1) AS pct
FROM io_inp_claims i
JOIN io_analytic a ON i.DSYSRTKY = a.DSYSRTKY
CROSS JOIN total t
WHERE TRY_STRPTIME(i.ADMSN_DT, '%Y%m%d') BETWEEN (a.death_dt - INTERVAL 30 DAY) AND a.death_dt
  AND i.PRNCPAL_DGNS_CD IS NOT NULL AND i.PRNCPAL_DGNS_CD <> ''
GROUP BY LEFT(i.PRNCPAL_DGNS_CD, 3)
ORDER BY n DESC
LIMIT 10;
"""

df_full  = con.execute(sql_full).df()
df_3     = con.execute(sql_3char).df()
total_n  = con.execute("""
    SELECT COUNT(*) FROM io_inp_claims i
    JOIN io_analytic a ON i.DSYSRTKY = a.DSYSRTKY
    WHERE TRY_STRPTIME(i.ADMSN_DT, '%Y%m%d') BETWEEN (a.death_dt - INTERVAL 30 DAY) AND a.death_dt
      AND i.PRNCPAL_DGNS_CD IS NOT NULL AND i.PRNCPAL_DGNS_CD <> ''
""").fetchone()[0]
con.close()

# Enrich full codes
df_full["description"] = df_full["icd_code"].map(DESCRIPTIONS_FULL).fillna("(see ICD-10 reference)")
df_full.insert(0, "rank", range(1, len(df_full) + 1))
df_full.columns = ["Rank", "ICD-10 Code", "Diagnosis", "N", "%"]

# Enrich 3-char codes
df_3["description"] = df_3["icd3"].map(DESCRIPTIONS_3).fillna("(see ICD-10 reference)")
df_3.insert(0, "rank", range(1, len(df_3) + 1))
df_3.columns = ["Rank", "ICD-10 Category", "Diagnosis", "N", "%"]

print("\nSheet 1 — Full ICD-10 codes:")
print(df_full.to_string(index=False))
print("\nSheet 2 — Grouped by 3-character category:")
print(df_3.to_string(index=False))


# ── Helper: write a table to a worksheet ──────────────────────────────────────
def write_table(ws, df, title, subtitle, col_widths, total_n):
    # Title
    ncols = len(df.columns)
    last_col = openpyxl.utils.get_column_letter(ncols)
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws["A1"]
    tc.value = title
    tc.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    tc.fill = PatternFill("solid", fgColor=SCARLET)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells(f"A2:{last_col}2")
    sc = ws["A2"]
    sc.value = subtitle
    sc.font = Font(name="Calibri", italic=True, size=10, color="555555")
    sc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Header row
    for col, h in enumerate(df.columns, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=SCARLET_D40)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="thin", color=WHITE),
                             right=Side(style="thin", color=WHITE))
    ws.row_dimensions[3].height = 20

    # Data rows
    aligns = ["center"] + ["center", "left"] + ["center"] * (ncols - 2)
    for i, row in df.iterrows():
        r = 4 + i
        fill = PatternFill("solid", fgColor=LIGHT_GRAY if i % 2 == 1 else WHITE)
        for col, (val, align) in enumerate(zip(row, aligns), 1):
            if col == ncols:
                val = f"{val}%"
            elif col == ncols - 1:
                val = int(val)
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[r].height = 18

    # Footnote
    fn_row = 4 + len(df) + 1
    ws.merge_cells(f"A{fn_row}:{last_col}{fn_row}")
    fn = ws.cell(row=fn_row, column=1,
                 value="Admissions defined by principal diagnosis on inpatient claim with admission date "
                       "within 30 days prior to date of death. "
                       "Percentages reflect share of all qualifying admissions (not patients).")
    fn.font = Font(name="Calibri", size=9, italic=True, color="666666")
    fn.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[fn_row].height = 28

    total_row = fn_row + 1
    ws.merge_cells(f"A{total_row}:{last_col}{total_row}")
    tr = ws.cell(row=total_row, column=1, value=f"Total qualifying admissions (within 30 days of death): {total_n:,}")
    tr.font = Font(name="Calibri", size=9, bold=True, color="333333")
    tr.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[total_row].height = 16

    # Column widths
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


# ── Build workbook ─────────────────────────────────────────────────────────────
SUBTITLE = "HNC Patients Receiving Immune Checkpoint Inhibitors (N = 2,527) | Medicare 2017-2023"
wb = openpyxl.Workbook()

ws1 = wb.active
ws1.title = "All ICD-10 Codes"
write_table(ws1, df_full,
            "All Principal Diagnoses: Inpatient Admissions Within 30 Days of Death",
            SUBTITLE,
            {"A": 6, "B": 14, "C": 50, "D": 8, "E": 8},
            total_n)

ws2 = wb.create_sheet("By ICD-10 Category")
write_table(ws2, df_3,
            "Top 10 Diagnosis Categories: Inpatient Admissions Within 30 Days of Death",
            SUBTITLE,
            {"A": 6, "B": 16, "C": 50, "D": 8, "E": 8},
            total_n)

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
