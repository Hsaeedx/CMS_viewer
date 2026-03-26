"""
make_outcomes_table.py
Exports OR results to a formatted Excel workbook: F:\\CMS\\outcomes_tables.xlsx
Two comparisons × three outcomes → six sheets:
  "A - Dysphagia", "A - G-tube", "A - Tracheostomy"
  "B - Dysphagia", "B - G-tube", "B - Tracheostomy"
Rows = subgroups; columns = time points.
"""

import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import pandas as pd
import numpy as np
from scipy.stats import chi2_contingency
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"F:\CMS\outcomes_tables.xlsx"

COMPARISONS = [
    ('A', 'psm_matched_A', 'TORS alone', 'RT alone'),
    ('B', 'psm_matched_B', 'TORS + RT',  'CT/CRT'),
]

C77_CTE = """
    c77 AS (
        SELECT DISTINCT o.DSYSRTKY
        FROM opscc_cohort o
        JOIN inp_claimsk_all i ON i.DSYSRTKY = o.DSYSRTKY,
        UNNEST([i.PRNCPAL_DGNS_CD, i.ADMTG_DGNS_CD,
                i.ICD_DGNS_CD1,  i.ICD_DGNS_CD2,  i.ICD_DGNS_CD3,
                i.ICD_DGNS_CD4,  i.ICD_DGNS_CD5,  i.ICD_DGNS_CD6,
                i.ICD_DGNS_CD7,  i.ICD_DGNS_CD8,  i.ICD_DGNS_CD9,
                i.ICD_DGNS_CD10, i.ICD_DGNS_CD11, i.ICD_DGNS_CD12,
                i.ICD_DGNS_CD13, i.ICD_DGNS_CD14, i.ICD_DGNS_CD15]) AS t(code)
        WHERE code LIKE 'C77%'
          AND TRY_STRPTIME(i.THRU_DT, '%Y%m%d')
                  BETWEEN o.first_hnc_date - INTERVAL 90 DAY
                      AND o.first_hnc_date + INTERVAL 90 DAY
        UNION ALL
        SELECT DISTINCT o.DSYSRTKY
        FROM opscc_cohort o
        JOIN out_claimsk_all oc ON oc.DSYSRTKY = o.DSYSRTKY,
        UNNEST([oc.PRNCPAL_DGNS_CD, oc.ICD_DGNS_CD1, oc.ICD_DGNS_CD2,
                oc.ICD_DGNS_CD3,  oc.ICD_DGNS_CD4, oc.ICD_DGNS_CD5,
                oc.ICD_DGNS_CD6,  oc.ICD_DGNS_CD7, oc.ICD_DGNS_CD8,
                oc.ICD_DGNS_CD9,  oc.ICD_DGNS_CD10]) AS t(code)
        WHERE code LIKE 'C77%'
          AND TRY_STRPTIME(oc.THRU_DT, '%Y%m%d')
                  BETWEEN o.first_hnc_date - INTERVAL 90 DAY
                      AND o.first_hnc_date + INTERVAL 90 DAY
        UNION ALL
        SELECT DISTINCT o.DSYSRTKY
        FROM opscc_cohort o
        JOIN car_linek_all cl ON cl.DSYSRTKY = o.DSYSRTKY
        WHERE cl.LINE_ICD_DGNS_CD LIKE 'C77%'
          AND TRY_STRPTIME(cl.THRU_DT, '%Y%m%d')
                  BETWEEN o.first_hnc_date - INTERVAL 90 DAY
                      AND o.first_hnc_date + INTERVAL 90 DAY
    )
"""

OUTCOMES = [
    ('Dysphagia',    'has_dysphagia',    'days_dys'),
    ('G-tube',       'has_gtube',        'days_gt'),
    ('Tracheostomy', 'has_tracheostomy', 'days_tr'),
]

TIMEPOINTS = [
    ('6-month',   182),
    ('1-year',    365),
    ('3-year',   1095),
    ('5-year',   1825),
    ('Anytime',   None),
]


def compute_or(sub, has_col, days_col, cutoff):
    valid = sub[sub[has_col].notna()].copy()
    if cutoff is None:
        elig = valid.copy()
        elig['ev'] = elig[has_col].astype(int)
    else:
        mask = (
            (valid['follow_up_days'] >= cutoff) |
            ((valid[has_col] == True) & (valid[days_col] <= cutoff))
        )
        elig = valid[mask].copy()
        elig['ev'] = ((elig[has_col] == True) & (elig[days_col] <= cutoff)).astype(int)

    t_ = elig[elig['tors'] == 1]
    c_ = elig[elig['tors'] == 0]
    nt, nc = len(t_), len(c_)
    et, ec = int(t_['ev'].sum()), int(c_['ev'].sum())
    pt = f"{100*et/nt:.1f}" if nt > 0 else "N/A"
    pc = f"{100*ec/nc:.1f}" if nc > 0 else "N/A"

    a, b, c, d = et, nt - et, ec, nc - ec
    if 0 in (a, b, c, d) or nt < 10 or nc < 10:
        return nt, et, pt, nc, ec, pc, np.nan, np.nan, np.nan, np.nan

    or_v   = (a * d) / (b * c)
    log_or = np.log(or_v)
    se     = np.sqrt(1/a + 1/b + 1/c + 1/d)
    lo, hi = np.exp(log_or - 1.96*se), np.exp(log_or + 1.96*se)
    _, p, _, _ = chi2_contingency([[a, b], [c, d]], correction=False)
    return nt, et, pt, nc, ec, pc, or_v, lo, hi, p


# ── Style helpers ──────────────────────────────────────────────────────────────
def border(style='thin'):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

def font(bold=False, size=10, color='000000', name='Calibri'):
    return Font(bold=bold, size=size, color=color, name=name)

def align(h='center', v='center', wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

HDR_FILL  = fill('BA0C2F')   # OSU Scarlet
SUB_FILL  = fill('70071C')   # Scarlet Dark
ALT_FILL  = fill('F2F2F2')
SIG_FILL  = fill('E2EFDA')
STRM_FILL = fill('F5D0D7')

# ── Load data for each comparison and build workbook ──────────────────────────
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12; SET temp_directory='F:\\CMS\\duckdb_temp';")

wb = openpyxl.Workbook()
wb.remove(wb.active)

N_TP = len(TIMEPOINTS)
M    = 4   # metrics per time point: TORS n/N, ctrl n/N, OR (95%CI), p
SCOL = 1
DCOL = 2

for comp, match_col, tors_label, ctrl_label in COMPARISONS:

    print(f"Loading data for Comparison {comp}: {tors_label} vs {ctrl_label}...")

    df = con.execute(f"""
        WITH {C77_CTE},
        matched AS (
            SELECT p.DSYSRTKY, p.tx_group, p.first_tx_date,
                   p.age_at_dx, p.van_walraven_score
            FROM opscc_propensity p
            WHERE p.{match_col} = TRUE
              AND p.tx_group IN ('{tors_label}', '{ctrl_label}')
              AND p.DSYSRTKY NOT IN (SELECT DSYSRTKY FROM c77)
        ),
        mbsf_sum AS (
            SELECT
                sub.DSYSRTKY,
                COALESCE(
                    MIN(CASE WHEN NOT (sub.buyin = '3' AND sub.hmoind IN ('0','4'))
                             THEN sub.mo_start END) - INTERVAL 1 DAY,
                    MAX(sub.mo_end)
                ) AS last_ffs_date,
                MAX(CASE WHEN sub.DEATH_DT IS NOT NULL AND sub.DEATH_DT != ''
                         THEN TRY_STRPTIME(sub.DEATH_DT, '%Y%m%d') END) AS death_date
            FROM (
                SELECT
                    m.DSYSRTKY, m.DEATH_DT,
                    make_date(CAST(m.RFRNC_YR AS INTEGER), t.mo, 1)      AS mo_start,
                    make_date(CAST(m.RFRNC_YR AS INTEGER), t.mo, 1)
                        + INTERVAL 1 MONTH - INTERVAL 1 DAY              AS mo_end,
                    CASE t.mo
                        WHEN 1  THEN m.BUYIN1  WHEN 2  THEN m.BUYIN2  WHEN 3  THEN m.BUYIN3
                        WHEN 4  THEN m.BUYIN4  WHEN 5  THEN m.BUYIN5  WHEN 6  THEN m.BUYIN6
                        WHEN 7  THEN m.BUYIN7  WHEN 8  THEN m.BUYIN8  WHEN 9  THEN m.BUYIN9
                        WHEN 10 THEN m.BUYIN10 WHEN 11 THEN m.BUYIN11 WHEN 12 THEN m.BUYIN12
                    END AS buyin,
                    CASE t.mo
                        WHEN 1  THEN m.HMOIND1  WHEN 2  THEN m.HMOIND2  WHEN 3  THEN m.HMOIND3
                        WHEN 4  THEN m.HMOIND4  WHEN 5  THEN m.HMOIND5  WHEN 6  THEN m.HMOIND6
                        WHEN 7  THEN m.HMOIND7  WHEN 8  THEN m.HMOIND8  WHEN 9  THEN m.HMOIND9
                        WHEN 10 THEN m.HMOIND10 WHEN 11 THEN m.HMOIND11 WHEN 12 THEN m.HMOIND12
                    END AS hmoind
                FROM mbsf_all m
                JOIN matched p ON m.DSYSRTKY = p.DSYSRTKY,
                UNNEST([1,2,3,4,5,6,7,8,9,10,11,12]) AS t(mo)
                WHERE make_date(CAST(m.RFRNC_YR AS INTEGER), t.mo, 1)
                          >= date_trunc('month', p.first_tx_date) + INTERVAL 1 MONTH
            ) sub
            GROUP BY sub.DSYSRTKY
        )
        SELECT
            p.DSYSRTKY, p.tx_group, p.first_tx_date,
            p.age_at_dx, p.van_walraven_score,
            s.death_date, s.last_ffs_date,
            o.has_dysphagia,
            DATEDIFF('day', p.first_tx_date, o.first_dysphagia_date) AS days_dys,
            o.has_gtube,
            DATEDIFF('day', p.first_tx_date, o.first_gtube_date)     AS days_gt,
            o.has_tracheostomy,
            DATEDIFF('day', p.first_tx_date, o.first_trach_date)     AS days_tr
        FROM matched p
        JOIN mbsf_sum       s ON s.DSYSRTKY = p.DSYSRTKY
        JOIN opscc_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
    """).df()

    df['death_date']     = pd.to_datetime(df['death_date'])
    df['last_ffs_date']  = pd.to_datetime(df['last_ffs_date'])
    df['first_tx_date']  = pd.to_datetime(df['first_tx_date'])
    df['censor_date']    = df['last_ffs_date']
    mask = df['death_date'].notna() & (df['death_date'] < df['last_ffs_date'])
    df.loc[mask, 'censor_date'] = df.loc[mask, 'death_date']
    df['follow_up_days'] = (df['censor_date'] - df['first_tx_date']).dt.days
    df['tors']           = (df['tx_group'] == tors_label).astype(int)

    q1 = df['van_walraven_score'].quantile(1/3)
    q2 = df['van_walraven_score'].quantile(2/3)
    df['elix_grp'] = pd.cut(df['van_walraven_score'],
                             bins=[-np.inf, q1, q2, np.inf],
                             labels=['Low', 'Mid', 'High'])

    strata = [
        ('All matched',                       df),
        ('Age < 75',                          df[df['age_at_dx'] < 75]),
        (f'Age \u2265 75',                    df[df['age_at_dx'] >= 75]),
        (f'Low comorbidity (VW\u2264{q1:.0f})',     df[df['elix_grp'] == 'Low']),
        (f'Mid comorbidity (VW {q1:.0f}\u2013{q2:.0f})', df[df['elix_grp'] == 'Mid']),
        (f'High comorbidity (VW>{q2:.0f})',   df[df['elix_grp'] == 'High']),
    ]

    for out_lbl, has_col, days_col in OUTCOMES:
        sheet_name = f"{comp} - {out_lbl}"
        ws = wb.create_sheet(title=sheet_name)
        total_cols = SCOL + N_TP * M

        # Row 1: title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1,
                    value=f"Comparison {comp}: {tors_label} vs {ctrl_label}  \u2014  "
                          f"{out_lbl}  \u2014  OR < 1 favors {tors_label}  (PSM, C77 excl.)")
        c.font      = font(bold=True, size=12)
        c.alignment = align(h='left', wrap=False)
        ws.row_dimensions[1].height = 22

        # Row 2-3: headers
        ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
        c = ws.cell(row=2, column=1, value='Subgroup')
        c.font = font(bold=True, size=11, color='FFFFFF')
        c.fill = HDR_FILL; c.alignment = align(); c.border = border()

        for tp_idx, (tp_lbl, _) in enumerate(TIMEPOINTS):
            sc = DCOL + tp_idx * M
            ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=sc + M - 1)
            c = ws.cell(row=2, column=sc, value=tp_lbl)
            c.font = font(bold=True, size=11, color='FFFFFF')
            c.fill = HDR_FILL; c.alignment = align(); c.border = border()
        ws.row_dimensions[2].height = 18

        tors_abbr = tors_label.replace(' ','')[:8]
        ctrl_abbr = ctrl_label.replace(' ','')[:8]
        sub_labels = [f'{tors_abbr}\nn/N (%)', f'{ctrl_abbr}\nn/N (%)', 'OR (95% CI)', 'p-value']
        for tp_idx in range(N_TP):
            sc = DCOL + tp_idx * M
            for j, lbl in enumerate(sub_labels):
                c = ws.cell(row=3, column=sc + j, value=lbl)
                c.font = font(bold=True, size=9); c.fill = SUB_FILL
                c.alignment = align(); c.border = border()
        ws.row_dimensions[3].height = 30

        # Data rows
        for row_i, (s_lbl, sub) in enumerate(strata):
            data_row = 4 + row_i
            row_fill = STRM_FILL if row_i == 0 else (ALT_FILL if row_i % 2 == 1 else fill('FFFFFF'))

            c = ws.cell(row=data_row, column=1, value=s_lbl)
            c.font      = font(bold=(row_i == 0), size=10)
            c.fill      = row_fill
            c.alignment = align(h='left', wrap=False)
            c.border    = border()

            for tp_idx, (_, cutoff) in enumerate(TIMEPOINTS):
                sc = DCOL + tp_idx * M
                nt, et, pt, nc, ec, pc, or_v, lo, hi, p = \
                    compute_or(sub, has_col, days_col, cutoff)

                sig  = (not np.isnan(or_v)) and (p < 0.05)
                cfil = SIG_FILL if sig else row_fill

                c = ws.cell(row=data_row, column=sc,     value=f"{et}/{nt}\n({pt}%)")
                c.font = font(size=10); c.fill = cfil; c.alignment = align(); c.border = border()

                c = ws.cell(row=data_row, column=sc + 1, value=f"{ec}/{nc}\n({pc}%)")
                c.font = font(size=10); c.fill = cfil; c.alignment = align(); c.border = border()

                or_str = 'N/A' if np.isnan(or_v) else f"{or_v:.2f}\n({lo:.2f}\u2013{hi:.2f})"
                c = ws.cell(row=data_row, column=sc + 2, value=or_str)
                c.font = font(bold=sig, size=10); c.fill = cfil; c.alignment = align(); c.border = border()

                if np.isnan(or_v):
                    p_str, p_col = '\u2014', '000000'
                elif p < 0.001:
                    p_str, p_col = '<0.001', ('CC0000' if sig else '000000')
                else:
                    p_str, p_col = f"{p:.3f}", ('CC0000' if sig else '000000')
                c = ws.cell(row=data_row, column=sc + 3, value=p_str)
                c.font = font(bold=sig, size=10, color=p_col)
                c.fill = cfil; c.alignment = align(); c.border = border()

            ws.row_dimensions[data_row].height = 32

        ws.column_dimensions['A'].width = 30
        for tp_idx in range(N_TP):
            sc = DCOL + tp_idx * M
            ws.column_dimensions[get_column_letter(sc)].width     = 11
            ws.column_dimensions[get_column_letter(sc + 1)].width = 11
            ws.column_dimensions[get_column_letter(sc + 2)].width = 14
            ws.column_dimensions[get_column_letter(sc + 3)].width = 8
        ws.freeze_panes = 'B4'

con.close()

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
