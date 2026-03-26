"""
make_table1.py
Cohort demographics for all four treatment groups, pre- and post-matching.

Columns:
  Pre-match: TORS alone | RT alone | TORS + RT | CT/CRT
  Comparison A post-match: TORS alone | RT alone | SMD
  Comparison B post-match: TORS + RT  | CT/CRT   | SMD

Output: F:\\CMS\\projects\\opscc\\table1_psm.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb, numpy as np, pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"F:\CMS\projects\opscc\table1_psm.xlsx"

FOUR_GROUPS = ['TORS alone', 'RT alone', 'TORS + RT', 'CT/CRT']

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        tx_group, psm_matched_A, psm_matched_B,
        age_at_dx, age_group, sex, race,
        van_walraven_score,
        chf, carit, valv, pcd, pvd, hypunc, hypc, para, ond, cpd,
        diabunc, diabc, hypothy, rf, ld, pud, aids, lymph, metacanc,
        solidtum, rheumd, coag, obes, wloss, fed, blane, dane,
        alcohol, drug, psycho, depre
    FROM opscc_propensity
    WHERE tx_group IN ('TORS alone', 'RT alone', 'TORS + RT', 'CT/CRT')
""").df()
con.close()

elix_flags = [
    'chf','carit','valv','pcd','pvd','hypunc','hypc','para','ond','cpd',
    'diabunc','diabc','hypothy','rf','ld','pud','aids','lymph','metacanc',
    'solidtum','rheumd','coag','obes','wloss','fed','blane','dane',
    'alcohol','drug','psycho','depre'
]
df[elix_flags]          = df[elix_flags].fillna(0).astype(int)
df['van_walraven_score'] = df['van_walraven_score'].fillna(0)
df['hypertension']       = ((df['hypunc'] == 1) | (df['hypc'] == 1)).astype(int)
df['diabetes']           = ((df['diabunc'] == 1) | (df['diabc'] == 1)).astype(int)
df['race_other']         = df['race'].isin(['Other', 'Unknown', 'AIAN']).astype(int)

# ── Pre-match groups ───────────────────────────────────────────────────────────
pre = {g: df[df['tx_group'] == g] for g in FOUR_GROUPS}

# ── Post-match groups ──────────────────────────────────────────────────────────
post_A = {g: df[(df['tx_group'] == g) & (df['psm_matched_A'] == True)]
           for g in ['TORS alone', 'RT alone']}
post_B = {g: df[(df['tx_group'] == g) & (df['psm_matched_B'] == True)]
           for g in ['TORS + RT', 'CT/CRT']}

for g in FOUR_GROUPS:
    print(f"  Pre-match {g}: {len(pre[g]):,}")
for g, d in post_A.items():
    print(f"  Post-match A {g}: {len(d):,}")
for g, d in post_B.items():
    print(f"  Post-match B {g}: {len(d):,}")

# ── SMD helpers ────────────────────────────────────────────────────────────────
def smd_cont(a, b):
    d = a.mean() - b.mean()
    p = np.sqrt((a.std()**2 + b.std()**2) / 2)
    return abs(d / p) if p > 0 else 0.0

def smd_bin(a, b):
    p1, p2 = a.mean(), b.mean()
    denom = np.sqrt((p1*(1-p1) + p2*(1-p2)) / 2)
    return abs((p1 - p2) / denom) if denom > 0 else 0.0

def mean_sd(s):
    return f"{s.mean():.1f} ({s.std():.1f})"

def pct(s, val=None):
    p = 100.0*(s == val).mean() if val is not None else 100.0*s.mean()
    n = int((s == val).sum()) if val is not None else int(s.sum())
    return f"{n:,} ({p:.1f}%)"

def fmt_smd(v):
    return f"{v:.3f}"

# ── Build table rows ───────────────────────────────────────────────────────────
rows = []

def section(title):
    rows.append({'label': title, 'is_section': True,
                 'c1':'','c2':'','c3':'','c4':'',
                 'a1':'','a2':'','smd_a':'',
                 'b1':'','b2':'','smd_b':''})

def add_row(label, vals, smd_a_val, smd_b_val, indent=False):
    rows.append({
        'label':   ('    ' + label) if indent else label,
        'is_section': False,
        'c1': vals[0], 'c2': vals[1], 'c3': vals[2], 'c4': vals[3],
        'a1': vals[4], 'a2': vals[5], 'smd_a': fmt_smd(smd_a_val),
        'b1': vals[6], 'b2': vals[7], 'smd_b': fmt_smd(smd_b_val),
    })

def cont(label, col, indent=False):
    vals = [
        mean_sd(pre['TORS alone'][col]),
        mean_sd(pre['RT alone'][col]),
        mean_sd(pre['TORS + RT'][col]),
        mean_sd(pre['CT/CRT'][col]),
        mean_sd(post_A['TORS alone'][col]),
        mean_sd(post_A['RT alone'][col]),
        mean_sd(post_B['TORS + RT'][col]),
        mean_sd(post_B['CT/CRT'][col]),
    ]
    smd_a = smd_cont(post_A['TORS alone'][col], post_A['RT alone'][col])
    smd_b = smd_cont(post_B['TORS + RT'][col],  post_B['CT/CRT'][col])
    add_row(label, vals, smd_a, smd_b, indent=indent)

def binary(label, col, val=None, indent=False):
    def p(s): return pct(s[col], val) if val is not None else pct(s[col])
    def s_bin(a, b):
        a2 = (a[col] == val).astype(float) if val is not None else a[col].astype(float)
        b2 = (b[col] == val).astype(float) if val is not None else b[col].astype(float)
        return smd_bin(a2, b2)
    vals = [
        p(pre['TORS alone']), p(pre['RT alone']),
        p(pre['TORS + RT']), p(pre['CT/CRT']),
        p(post_A['TORS alone']), p(post_A['RT alone']),
        p(post_B['TORS + RT']), p(post_B['CT/CRT']),
    ]
    smd_a = s_bin(post_A['TORS alone'], post_A['RT alone'])
    smd_b = s_bin(post_B['TORS + RT'],  post_B['CT/CRT'])
    add_row(label, vals, smd_a, smd_b, indent=indent)

section('DEMOGRAPHICS')
cont  ('Age at diagnosis, mean (SD)', 'age_at_dx')
binary('Male sex', 'sex', 'Male')

section('RACE / ETHNICITY')
binary('White',           'race', 'White',    indent=True)
binary('Black',           'race', 'Black',    indent=True)
binary('Hispanic',        'race', 'Hispanic', indent=True)
binary('Asian / PI',      'race', 'Asian/PI', indent=True)
binary('Other / Unknown', 'race_other',       indent=True)

section('COMORBIDITIES')
cont  ('van Walraven score, mean (SD)', 'van_walraven_score')
binary('Hypertension',             'hypertension', indent=True)
binary('Chronic pulmonary disease','cpd',          indent=True)
binary('Diabetes mellitus',        'diabetes',     indent=True)
binary('Cardiac arrhythmia',       'carit',        indent=True)
binary('Congestive heart failure', 'chf',          indent=True)
binary('Renal failure',            'rf',           indent=True)
binary('Alcohol abuse',            'alcohol',      indent=True)
binary('Depression',               'depre',        indent=True)
binary('Obesity',                  'obes',         indent=True)
binary('Other neurological',       'ond',          indent=True)

df_table = pd.DataFrame(rows)

# ── Write Excel ────────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='BA0C2F')   # OSU Scarlet
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='F5D0D7')
SECTION_FONT = Font(bold=True, size=10, color='70071C')
ALT_FILL     = PatternFill('solid', fgColor='FDF5F6')
SMD_WARN     = PatternFill('solid', fgColor='FFE699')
TITLE_FONT   = Font(bold=True, size=13, color='BA0C2F')
GROUP_FILL   = PatternFill('solid', fgColor='E8C0C8')
GROUP_FONT   = Font(bold=True, size=10, color='4A0513')

wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = 'Table 1'

ws.append(['Table 1. Cohort Characteristics — Pre- and Post-Propensity Score Matching'])
ws['A1'].font = TITLE_FONT
ws.append([])

nTA = len(pre['TORS alone']); nRA = len(pre['RT alone'])
nTR = len(pre['TORS + RT']);  nCC = len(pre['CT/CRT'])
nTA_m = len(post_A['TORS alone']); nRA_m = len(post_A['RT alone'])
nTR_m = len(post_B['TORS + RT']);  nCC_m = len(post_B['CT/CRT'])

hrow = ws.max_row + 1

# Group header row (row 3): Pre-match (cols 2-5) | Comp A post-match (6-8) | Comp B post-match (9-11)
groups = [
    (2, 5, 'Pre-Match (all patients)'),
    (6, 8, 'Comparison A — Post-Match\nTORS alone vs RT alone'),
    (9, 11, 'Comparison B — Post-Match\nTORS + RT vs CT/CRT'),
]
for sc, ec, label in groups:
    ws.merge_cells(start_row=hrow, start_column=sc, end_row=hrow, end_column=ec)
    c = ws.cell(row=hrow, column=sc, value=label)
    c.fill      = GROUP_FILL
    c.font      = GROUP_FONT
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[hrow].height = 30

# Column sub-headers (row 4)
srow = hrow + 1
col_headers = [
    'Characteristic',
    f'TORS alone\n(n={nTA:,})',
    f'RT alone\n(n={nRA:,})',
    f'TORS + RT\n(n={nTR:,})',
    f'CT/CRT\n(n={nCC:,})',
    f'TORS alone\n(n={nTA_m:,})',
    f'RT alone\n(n={nRA_m:,})',
    'SMD',
    f'TORS + RT\n(n={nTR_m:,})',
    f'CT/CRT\n(n={nCC_m:,})',
    'SMD',
]
for ci, cn in enumerate(col_headers, 1):
    cell = ws.cell(row=srow, column=ci, value=cn)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center' if ci > 1 else 'left',
                                wrap_text=True, vertical='center')
ws.row_dimensions[srow].height = 36

# Data rows
alt = 0
for ri, row_data in enumerate(df_table.itertuples(index=False), srow + 1):
    is_sec = row_data.is_section
    if not is_sec:
        alt += 1
    vals = [row_data.label,
            row_data.c1, row_data.c2, row_data.c3, row_data.c4,
            row_data.a1, row_data.a2, row_data.smd_a,
            row_data.b1, row_data.b2, row_data.smd_b]
    for ci, val in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if is_sec:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif alt % 2 == 0:
            cell.fill = ALT_FILL

        # Highlight imbalanced SMD
        if ci in (8, 11) and not is_sec and val != '':
            try:
                if float(val) >= 0.10:
                    cell.fill = SMD_WARN
            except (ValueError, TypeError):
                pass

        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center', wrap_text=(ci == 1))

# Column widths
ws.column_dimensions['A'].width = 40
for letter in ['B','C','D','E','F','G','H','I','J','K']:
    ws.column_dimensions[letter].width = 18

ws.freeze_panes = f'B{srow + 1}'

footer_row = ws.max_row + 2
ws.cell(row=footer_row, column=1,
        value='SMD = standardized mean difference (values \u22650.10 highlighted). '
              'Continuous variables: mean (SD). Categorical: n (%). '
              'Elixhauser comorbidities: 12-month lookback before diagnosis. '
              'Comp A = TORS alone vs RT alone; Comp B = TORS + RT vs CT/CRT.')
ws.cell(row=footer_row, column=1).font = Font(italic=True, size=8, color='555555')

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
