"""
make_gradient_elix.py

1. dose_response_gradient.png — HR by SLP timing group showing dose-response
2. elix_stratified_results.xlsx — Cox PH stratified by Elixhauser (van Walraven) quartile
"""
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
load_dotenv(Path(__file__).resolve().parents[2] / ".env")

import duckdb
import numpy as np
import pandas as pd
from lifelines import CoxPHFitter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_out_dir      = Path(os.getenv("project_paths", ".")) / "stroke_SLP"
_out_dir.mkdir(parents=True, exist_ok=True)
DB_PATH       = Path(os.getenv("duckdb_database", "cms_data.duckdb"))
GRADIENT_PATH = _out_dir / "dose_response_gradient.png"
ELIX_PATH     = _out_dir / "elix_stratified_results.xlsx"
LANDMARK      = 90
MAX_FOLLOW_LM = 1795

# ── Load & filter ─────────────────────────────────────────────────────────────
print("Loading data...")
con = duckdb.connect(str(DB_PATH), read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT p.DSYSRTKY, p.age_at_adm, p.sex, p.stroke_type,
           p.index_los, p.mech_vent, p.peg_placed, p.trach_placed,
           p.dschg_status, p.dysphagia_poa, p.aspiration_poa,
           p.prior_stroke, p.dementia, p.van_walraven_score,
           p.slp_outpt_any_90d, p.days_to_slp_outpt,
           o.days_to_death, o.days_to_aspiration,
           o.days_to_gtube, o.pre_stroke_tube
    FROM stroke_propensity p
    JOIN stroke_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
""").df()
con.close()
print(f"  Loaded {len(df):,} rows")

lm = df[df['days_to_death'].isna() | (df['days_to_death'] > LANDMARK)].copy()
lm_slp = lm[lm['slp_outpt_any_90d'] == 1].copy()

def slp_group(d):
    if pd.isna(d): return None
    if d <= 14:    return 'SLP 0-14d'
    if d <= 30:    return 'SLP 15-30d'
    return 'SLP 31-90d'

lm_slp['timing_group'] = lm_slp['days_to_slp_outpt'].apply(slp_group)

def map_dschg(c):
    if pd.isna(c): return 'home'
    c = str(c).strip()
    if c in ('01','08'): return 'home'
    if c == '06':        return 'hha'
    if c in ('03','61'): return 'snf'
    if c == '62':        return 'irf'
    return 'other'

lm_slp['dschg_group'] = lm_slp['dschg_status'].apply(map_dschg)
lm_slp = lm_slp[lm_slp['dschg_group'].isin(['home','hha'])].copy()
print(f"  After home/HHA filter: {len(lm_slp):,}")

# Landmark-shifted times
for col in ['days_to_death','days_to_aspiration','days_to_gtube']:
    lm_slp[col] = lm_slp[col].astype('float64')
    lm_slp[f'{col}_lm'] = np.where(
        lm_slp[col].notna() & (lm_slp[col] > LANDMARK),
        (lm_slp[col] - LANDMARK).clip(lower=0.5), np.nan)

lm_slp['censor_lm'] = np.where(
    lm_slp['days_to_death_lm'].notna(),
    lm_slp['days_to_death_lm'].clip(upper=MAX_FOLLOW_LM), MAX_FOLLOW_LM)

lm_slp['died_lm']  = lm_slp['days_to_death_lm'].notna().astype(int)
lm_slp['asp_lm']   = lm_slp['days_to_aspiration_lm'].notna().astype(int)
lm_slp['gtube_lm'] = lm_slp['days_to_gtube_lm'].notna().astype(int)

lm_slp['days_to_aspiration_lm_filled'] = (
    lm_slp['days_to_aspiration_lm'].fillna(lm_slp['censor_lm']).clip(lower=0.5))
lm_slp['days_to_gtube_lm_filled'] = (
    lm_slp['days_to_gtube_lm'].fillna(lm_slp['censor_lm']).clip(lower=0.5))

# Covariates
lm_slp['slp_15_30d'] = (lm_slp['timing_group'] == 'SLP 15-30d').astype(int)
lm_slp['slp_31_90d'] = (lm_slp['timing_group'] == 'SLP 31-90d').astype(int)
# Ordinal trend variable: 1=0-14d, 2=15-30d, 3=31-90d
lm_slp['slp_ordinal'] = lm_slp['timing_group'].map(
    {'SLP 0-14d': 1, 'SLP 15-30d': 2, 'SLP 31-90d': 3})

dschg_d  = pd.get_dummies(lm_slp['dschg_group'], prefix='dschg', drop_first=True)
sex_d    = pd.get_dummies(lm_slp['sex'],          prefix='sex',   drop_first=True)
stroke_d = pd.get_dummies(lm_slp['stroke_type'],  prefix='stroke',drop_first=True)
lm_slp   = pd.concat([lm_slp, dschg_d, sex_d, stroke_d], axis=1)

_cands = (['slp_15_30d','slp_31_90d',
           'age_at_adm','van_walraven_score','index_los',
           'mech_vent','trach_placed','prior_stroke','dementia',
           'dysphagia_poa','aspiration_poa']
          + list(dschg_d.columns) + list(sex_d.columns) + list(stroke_d.columns))
COVS = [c for c in _cands if c in lm_slp.columns and lm_slp[c].std() > 0.05]

# Trend covariates (replace slp_15_30d/slp_31_90d with ordinal)
COVS_TREND = ['slp_ordinal'] + [c for c in COVS
                                  if c not in ('slp_15_30d','slp_31_90d')]

MODELS = [
    ('Aspiration PNA', 'asp_lm',   'days_to_aspiration_lm_filled', False, 0.0),
    ('PEG/G-tube',     'gtube_lm', 'days_to_gtube_lm_filled',      True,  0.1),
    ('Mortality',      'died_lm',  'censor_lm',                    False, 0.0),
]


# ── Helper: run one Cox model, return HR rows ──────────────────────────────────
def run_cox(sub, dur_col, ev_col, cov_list, pen, terms):
    """Returns dict {term: (hr, lo, hi, p)} or None on failure."""
    cox_df = sub[[dur_col, ev_col] + cov_list].rename(
        columns={dur_col: 'duration', ev_col: 'event'}).dropna()
    if cox_df['event'].sum() < 10:
        return None, len(cox_df), int(cox_df['event'].sum())
    try:
        cph = CoxPHFitter(penalizer=pen)
        cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
        out = {}
        for t in terms:
            if t not in cph.summary.index:
                continue
            r = cph.summary.loc[t]
            out[t] = (np.exp(r['coef']),
                      np.exp(r['coef lower 95%']),
                      np.exp(r['coef upper 95%']),
                      r['p'])
        return out, len(cox_df), int(cox_df['event'].sum())
    except Exception as e:
        print(f"    Cox error: {e}")
        return None, 0, 0


# ═══════════════════════════════════════════════════════════════════════════════
# DOSE-RESPONSE GRADIENT
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding dose-response gradient...")

# Compute median days per group (for x-axis)
med_days = lm_slp.groupby('timing_group')['days_to_slp_outpt'].median()
X_DAYS = {
    'SLP 0-14d':  med_days.get('SLP 0-14d',   7),
    'SLP 15-30d': med_days.get('SLP 15-30d', 22),
    'SLP 31-90d': med_days.get('SLP 31-90d', 52),
}
x_vals = [X_DAYS['SLP 0-14d'], X_DAYS['SLP 15-30d'], X_DAYS['SLP 31-90d']]
print(f"  Median days: {X_DAYS}")

# Collect HRs for gradient
gradient_data = {}   # outcome -> {'x': [...], 'hr': [...], 'lo': [...], 'hi': [...], 'p_trend': p}

for label, ev_col, dur_col, excl_peg, pen in MODELS:
    sub = lm_slp.copy()
    if excl_peg:
        sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
    cov_use = list(COVS)
    if not excl_peg and 'peg_placed' not in cov_use:
        cov_use = ['peg_placed'] + cov_use

    # Main model (binary dummies) for HR at 15-30d and 31-90d
    res, n, n_ev = run_cox(sub, dur_col, ev_col, cov_use, pen,
                           ['slp_15_30d', 'slp_31_90d'])

    # Trend model (ordinal)
    cov_trend = ['slp_ordinal'] + [c for c in cov_use
                                    if c not in ('slp_15_30d','slp_31_90d')]
    res_trend, _, _ = run_cox(sub, dur_col, ev_col, cov_trend, pen, ['slp_ordinal'])

    if res is None:
        continue

    hr_15, lo_15, hi_15, _ = res.get('slp_15_30d', (None,)*4)
    hr_31, lo_31, hi_31, _ = res.get('slp_31_90d', (None,)*4)
    p_trend = res_trend['slp_ordinal'][3] if (res_trend and 'slp_ordinal' in res_trend) else np.nan

    gradient_data[label] = {
        'x':       x_vals,
        'hr':      [1.0,    hr_15,  hr_31],
        'lo':      [np.nan, lo_15,  lo_31],
        'hi':      [np.nan, hi_15,  hi_31],
        'n':       n,
        'events':  n_ev,
        'p_trend': p_trend,
    }
    print(f"  {label}: HR 15-30d={hr_15:.2f}, HR 31-90d={hr_31:.2f}, p-trend={p_trend:.4f}")

# ── Plot ──────────────────────────────────────────────────────────────────────
OUTCOME_COLORS  = {
    'Aspiration PNA': '#1F4E79',
    'PEG/G-tube':     '#C55A11',
    'Mortality':      '#375623',
}
OUTCOME_MARKERS = {
    'Aspiration PNA': 'o',
    'PEG/G-tube':     's',
    'Mortality':      '^',
}

fig, ax = plt.subplots(figsize=(9, 6), facecolor='white')
ax.set_facecolor('white')

ax.axhline(1.0, color='#888888', lw=1.2, linestyle='--', zorder=1, label='_nolegend_')
ax.axhspan(0.95, 1.05, color='#EEEEEE', alpha=0.5, zorder=0)  # band around null

for label, d in gradient_data.items():
    col = OUTCOME_COLORS[label]
    mkr = OUTCOME_MARKERS[label]
    x   = d['x']
    hr  = d['hr']
    lo  = d['lo']
    hi  = d['hi']
    p_t = d['p_trend']
    p_str = 'p-trend <0.0001' if p_t < 0.0001 else f'p-trend = {p_t:.4f}'

    # Connecting line
    ax.plot(x, hr, color=col, lw=2.0, zorder=3, alpha=0.85)

    # CI bars (not for reference group)
    for xi, hri, loi, hii in zip(x, hr, lo, hi):
        if not np.isnan(loi):
            ax.plot([xi, xi], [loi, hii], color=col, lw=1.4, zorder=3)
            ax.plot([xi-0.5, xi+0.5], [loi, loi], color=col, lw=1.0, zorder=3)
            ax.plot([xi-0.5, xi+0.5], [hii, hii], color=col, lw=1.0, zorder=3)

    # Dots
    ax.plot(x[0], hr[0], marker=mkr, color='white', markersize=10,
            markeredgecolor=col, markeredgewidth=2.0, zorder=4)   # reference: open
    for xi, hri in zip(x[1:], hr[1:]):
        ax.plot(xi, hri, marker=mkr, color=col, markersize=10,
                markeredgecolor='white', markeredgewidth=0.8, zorder=4)

    # p-trend label at end of line
    ax.text(x[-1] + 1.5, hr[-1],
            f"{label}\n({p_str})",
            va='center', ha='left', fontsize=9, color=col, fontweight='bold')

# X axis: median days
ax.set_xlim(0, 80)
ax.set_xticks(x_vals)
ax.set_xticklabels([
    f"SLP 0–14d\n(median {x_vals[0]:.0f}d)",
    f"SLP 15–30d\n(median {x_vals[1]:.0f}d)",
    f"SLP 31–90d\n(median {x_vals[2]:.0f}d)",
], fontsize=10)
ax.set_xlabel('Days from Discharge to First Outpatient SLP', fontsize=11, labelpad=8)

ax.set_yscale('log')
ax.set_ylim(0.70, 2.5)
ax.set_yticks([0.75, 1.0, 1.25, 1.5, 2.0])
ax.set_yticklabels(['0.75','1.00','1.25','1.50','2.00'], fontsize=10)
ax.set_ylabel('Adjusted Hazard Ratio (log scale)\nReference: SLP 0–14 days', fontsize=10.5)

# Reference group annotation
ax.text(x_vals[0], 1.03, 'Reference\n(HR = 1.00)', ha='center', va='bottom',
        fontsize=8.5, color='#555555', fontstyle='italic')

ax.set_title(
    'Dose–Response Gradient: Later SLP Initiation and Worse Post-Stroke Outcomes\n'
    'Home / HHA-discharged Medicare stroke patients  •  90-day landmark  •  N = 108,695',
    fontsize=11, fontweight='bold', color='#1F4E79', pad=10)

for sp in ['top','right']:
    ax.spines[sp].set_visible(False)
ax.spines['left'].set_color('#AAAAAA')
ax.spines['bottom'].set_color('#AAAAAA')
ax.tick_params(axis='both', color='#AAAAAA')

plt.tight_layout()
plt.savefig(str(GRADIENT_PATH), dpi=600, bbox_inches='tight', facecolor='white')
plt.close()
print(f"Saved: {GRADIENT_PATH}")


# ═══════════════════════════════════════════════════════════════════════════════
# ELIXHAUSER STRATIFIED RESULTS TABLE
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding Elixhauser stratified results...")

vw = lm_slp['van_walraven_score']
q1, q2, q3 = vw.quantile([0.25, 0.50, 0.75])
ql_labels = [
    f'Q1  (VW \u2264 {q1:.0f})  — Lowest comorbidity',
    f'Q2  (VW {q1:.0f}\u2013{q2:.0f})',
    f'Q3  (VW {q2:.0f}\u2013{q3:.0f})',
    f'Q4  (VW > {q3:.0f})  — Highest comorbidity',
]

def vw_quartile(score):
    if pd.isna(score): return None
    if score <= q1: return ql_labels[0]
    if score <= q2: return ql_labels[1]
    if score <= q3: return ql_labels[2]
    return ql_labels[3]

lm_slp['vw_q'] = lm_slp['van_walraven_score'].apply(vw_quartile)

elix_rows = []
for qlab in ql_labels:
    q_sub = lm_slp[lm_slp['vw_q'] == qlab]
    n_q   = len(q_sub)
    n_grp = {g: int((q_sub['timing_group']==g).sum())
              for g in ['SLP 0-14d','SLP 15-30d','SLP 31-90d']}
    print(f"  Q N={n_q:,}  {n_grp}")

    for label, ev_col, dur_col, excl_peg, pen in MODELS:
        sub = q_sub.copy()
        if excl_peg:
            sub = sub[(sub['peg_placed']==0) & (sub['pre_stroke_tube'].fillna(0)==0)]
        cov_use = list(COVS)
        if not excl_peg and 'peg_placed' not in cov_use:
            cov_use = ['peg_placed'] + cov_use

        res, n, n_ev = run_cox(sub, dur_col, ev_col, cov_use, pen,
                               ['slp_15_30d','slp_31_90d'])

        for comp, tv in [('15-30d vs 0-14d','slp_15_30d'),
                         ('31-90d vs 0-14d','slp_31_90d')]:
            base = {
                'Elixhauser Quartile': qlab,
                'N in quartile': f"{n_q:,}",
                'N (0-14d)': f"{n_grp['SLP 0-14d']:,}",
                'N (15-30d)': f"{n_grp['SLP 15-30d']:,}",
                'N (31-90d)': f"{n_grp['SLP 31-90d']:,}",
                'Outcome': label,
                'Comparison': comp,
                'N in model': f"{n:,}",
                'Events': f"{n_ev:,}",
            }
            if res is None or tv not in res:
                elix_rows.append({**base, 'HR (95% CI)': 'Too few events', 'p-value': ''})
                continue
            hr, lo, hi, p = res[tv]
            p_str = '<0.0001' if p < 0.0001 else f'{p:.4f}'
            elix_rows.append({**base,
                'HR (95% CI)': f"{hr:.2f}  [{lo:.2f}\u2013{hi:.2f}]",
                'p-value':     p_str,
            })

df_elix = pd.DataFrame(elix_rows)

# ── Write Excel ────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
QHDR_FILL   = PatternFill('solid', fgColor='2E5090')
QHDR_FONT   = Font(bold=True, color='FFFFFF', size=10)
OUT_FILL    = PatternFill('solid', fgColor='BDD7EE')
OUT_FONT    = Font(bold=True, size=10, color='1F4E79')
ALT_FILL    = PatternFill('solid', fgColor='EBF3FB')
TITLE_FONT  = Font(bold=True, size=12, color='1F4E79')
SIG_FONT    = Font(bold=True, size=10, color='C55A11')

DISPLAY_COLS = ['Elixhauser Quartile', 'Outcome', 'Comparison',
                'N in quartile', 'N (0-14d)', 'N (15-30d)', 'N (31-90d)',
                'N in model', 'Events', 'HR (95% CI)', 'p-value']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Elix_Stratified'

ws.append(['Table: Cox PH Results Stratified by Elixhauser Comorbidity (van Walraven Score)'])
ws['A1'].font = TITLE_FONT
ws.append([f'Reference: SLP 0\u201314d  \u2022  Quartile cutpoints: Q1\u2264{q1:.0f}, Q2\u2264{q2:.0f}, Q3\u2264{q3:.0f}  \u2022  90-day landmark  \u2022  Home/HHA only'])
ws['A2'].font = Font(italic=True, size=10, color='555555')
ws.append([])

hdr_row = ws.max_row + 1
for ci, col in enumerate(DISPLAY_COLS, 1):
    cell = ws.cell(row=hdr_row, column=ci, value=col)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[hdr_row].height = 30

prev_q = None
prev_out = None
alt = 0

for _, row in df_elix.iterrows():
    ri = ws.max_row + 1
    q   = row['Elixhauser Quartile']
    out = row['Outcome']

    # Quartile section header when quartile changes
    if q != prev_q:
        ws.row_dimensions[ri].height = 20
        for ci, col in enumerate(DISPLAY_COLS, 1):
            val = q if ci == 1 else ''
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = QHDR_FONT
            cell.fill      = QHDR_FILL
            cell.alignment = Alignment(horizontal='left' if ci==1 else 'center',
                                        vertical='center')
        prev_q  = q
        prev_out = None
        alt = 0
        ri = ws.max_row + 1

    # Outcome sub-header when outcome changes within a quartile
    if out != prev_out:
        ws.row_dimensions[ri].height = 16
        for ci, col in enumerate(DISPLAY_COLS, 1):
            val = out if ci == 2 else (row.get(col,'') if ci > 2 and col in
                  ['N in quartile','N (0-14d)','N (15-30d)','N (31-90d)'] else '')
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = OUT_FONT
            cell.fill      = OUT_FILL
            cell.alignment = Alignment(horizontal='left' if ci<=2 else 'center',
                                        vertical='center')
        prev_out = out
        alt = 0
        ri = ws.max_row + 1

    # Data row
    alt += 1
    ws.row_dimensions[ri].height = 16
    for ci, col in enumerate(DISPLAY_COLS, 1):
        # Don't repeat quartile/N info in data rows
        if col in ('Elixhauser Quartile','N in quartile',
                   'N (0-14d)','N (15-30d)','N (31-90d)','Outcome'):
            val = ''
        else:
            val = row.get(col, '')
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = Alignment(
            horizontal='left' if ci <= 2 else 'center',
            vertical='center')
        if alt % 2 == 0:
            cell.fill = ALT_FILL
        if col == 'p-value' and val == '<0.0001':
            cell.font = SIG_FONT
        if col == 'HR (95% CI)' and val not in ('', 'Too few events'):
            cell.font = Font(bold=True, size=10)

# Column widths
col_widths = [38, 16, 18, 12, 10, 10, 10, 12, 10, 22, 10]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Footnote
fn_row = ws.max_row + 2
ws.cell(row=fn_row, column=1,
        value=('Covariates in each stratum: age, LOS, SLP timing dummies, mechanical ventilation, '
               'tracheostomy, prior stroke, dementia, dysphagia POA, aspiration POA, '
               'discharge destination (home vs HHA), sex, stroke type.  '
               'Van Walraven score is a weighted Elixhauser comorbidity index.'))
ws.cell(row=fn_row, column=1).font = Font(italic=True, size=8.5, color='666666')
ws.merge_cells(start_row=fn_row, start_column=1,
               end_row=fn_row,   end_column=len(DISPLAY_COLS))

ws.freeze_panes = f'A{hdr_row + 1}'
wb.save(str(ELIX_PATH))
print(f"Saved: {ELIX_PATH}")
