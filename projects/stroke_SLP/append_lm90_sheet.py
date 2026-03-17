"""Appends LM8_LM90 sheet (90-day landmark sensitivity) to the existing Excel."""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')
import duckdb, numpy as np, pandas as pd
from lifelines import CoxPHFitter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT_PATH = r"F:\CMS\projects\stroke_SLP\stroke_landmark_tables_home_hha.xlsx"
LANDMARK, MAX_FOLLOW_LM = 90, 1795

con = duckdb.connect(r'F:\CMS\cms_data.duckdb', read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT p.DSYSRTKY, p.age_at_adm, p.sex, p.stroke_type,
           p.index_los, p.mech_vent, p.peg_placed, p.trach_placed,
           p.dschg_status, p.dysphagia_poa, p.aspiration_poa,
           p.prior_stroke, p.dementia, p.van_walraven_score,
           p.slp_outpt_any_90d, p.days_to_slp_outpt,
           o.days_to_death, o.days_to_aspiration,
           o.days_to_gtube, o.days_to_recur_stroke,
           o.pre_stroke_tube
    FROM stroke_propensity p
    JOIN stroke_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
""").df()
con.close()

lm = df[df['days_to_death'].isna() | (df['days_to_death'] > LANDMARK)].copy()
lm_slp = lm[lm['slp_outpt_any_90d'] == 1].copy()

def slp_group(d):
    if pd.isna(d): return None
    if d <= 14: return '0-14d'
    if d <= 30: return '15-30d'
    return '31-90d'

lm_slp['grp'] = lm_slp['days_to_slp_outpt'].apply(slp_group)

def map_dschg(c):
    c = str(c).strip() if not pd.isna(c) else ''
    if c in ('01','08'): return 'home'
    if c == '06': return 'hha'
    if c in ('03','61'): return 'snf'
    if c == '62': return 'irf'
    return 'other'

lm_slp['dschg_group'] = lm_slp['dschg_status'].apply(map_dschg)
lm_slp = lm_slp[lm_slp['dschg_group'].isin(['home','hha'])].copy()

for col in ['days_to_death','days_to_aspiration','days_to_gtube','days_to_recur_stroke']:
    lm_slp[col] = lm_slp[col].astype('float64')
    lm_slp[col+'_lm'] = np.where(
        lm_slp[col].notna() & (lm_slp[col] > LANDMARK),
        (lm_slp[col] - LANDMARK).clip(lower=0.5), np.nan)

lm_slp['censor_lm'] = np.where(
    lm_slp['days_to_death_lm'].notna(),
    lm_slp['days_to_death_lm'].clip(upper=MAX_FOLLOW_LM), MAX_FOLLOW_LM)
lm_slp['died_lm']  = lm_slp['days_to_death_lm'].notna().astype(int)
lm_slp['asp_lm']   = lm_slp['days_to_aspiration_lm'].notna().astype(int)
lm_slp['gtube_lm'] = lm_slp['days_to_gtube_lm'].notna().astype(int)
lm_slp['recur_lm'] = lm_slp['days_to_recur_stroke_lm'].notna().astype(int)

for ev, dur in [('asp_lm','days_to_aspiration_lm'),
                ('gtube_lm','days_to_gtube_lm'),
                ('recur_lm','days_to_recur_stroke_lm')]:
    lm_slp[dur+'_filled'] = lm_slp[dur].fillna(lm_slp['censor_lm']).clip(lower=0.5)

lm_slp['slp_15_30d'] = (lm_slp['grp'] == '15-30d').astype(int)
lm_slp['slp_31_90d'] = (lm_slp['grp'] == '31-90d').astype(int)

sex_d   = pd.get_dummies(lm_slp['sex'],         prefix='sex',    drop_first=True)
strk_d  = pd.get_dummies(lm_slp['stroke_type'], prefix='stroke', drop_first=True)
dschg_d = pd.get_dummies(lm_slp['dschg_group'], prefix='dschg',  drop_first=True)
lm_slp  = pd.concat([lm_slp, sex_d, strk_d, dschg_d], axis=1)

cand_covs = (['slp_15_30d','slp_31_90d','age_at_adm','van_walraven_score','index_los',
               'mech_vent','trach_placed','peg_placed','prior_stroke','dementia',
               'dysphagia_poa','aspiration_poa']
              + list(sex_d.columns) + list(strk_d.columns) + list(dschg_d.columns))
BASE_COVS = [c for c in cand_covs if c in lm_slp.columns and lm_slp[c].std() > 0.05]

MODELS = [
    ('Aspiration PNA', 'asp_lm',   'days_to_aspiration_lm_filled',  False, 0.0),
    ('PEG/G-tube',     'gtube_lm', 'days_to_gtube_lm_filled',       True,  0.1),
    ('Mortality',      'died_lm',  'censor_lm',                     False, 0.0),
    ('Recur stroke',   'recur_lm', 'days_to_recur_stroke_lm_filled', False, 0.0),
]

records = []
for label, ev_col, dur_col, excl_peg, pen in MODELS:
    sub = lm_slp.copy()
    if excl_peg:
        sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
    cov_use = [c for c in BASE_COVS if c != 'peg_placed' or not excl_peg]
    if not excl_peg and 'peg_placed' not in cov_use:
        cov_use = ['peg_placed'] + cov_use

    raw = sub[[dur_col, ev_col] + cov_use].rename(
        columns={dur_col: 'duration', ev_col: 'event'}).dropna()
    cov_ok = [c for c in cov_use if raw[c].std() > 0.05]
    cox_df = raw[['duration', 'event'] + cov_ok]
    n_ev   = int(cox_df['event'].sum())

    try:
        cph = CoxPHFitter(penalizer=pen)
        cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
        for tv, comp in [('slp_15_30d','15-30d vs 0-14d'),('slp_31_90d','31-90d vs 0-14d')]:
            if tv not in cph.summary.index: continue
            r  = cph.summary.loc[tv]
            hr = np.exp(r['coef']); lo = np.exp(r['coef lower 95%']); hi = np.exp(r['coef upper 95%'])
            p  = r['p']
            ev_v = hr + np.sqrt(hr * abs(hr-1))
            records.append({
                'Outcome':    label,
                'Comparison': comp,
                'N':          len(cox_df),
                'Events':     n_ev,
                'Event %':    round(100*n_ev/len(cox_df), 1),
                'HR':         round(hr, 2),
                '95% CI':     f"[{lo:.2f}, {hi:.2f}]",
                'p-value':    '<0.0001' if p < 0.0001 else f'{p:.4f}',
                'E-value':    round(ev_v, 2),
                'Note': 'Clock starts day 90 — all groups have started SLP by landmark',
            })
    except Exception as e:
        records.append({'Outcome': label, 'Comparison': 'ERROR', 'N': 0, 'Events': 0,
                        'Event %': 0, 'HR': None, '95% CI': str(e)[:60],
                        'p-value': '', 'E-value': None, 'Note': ''})

df_out = pd.DataFrame(records)

# Append to existing Excel
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL    = PatternFill('solid', fgColor='D6E4F0')
TITLE_FONT  = Font(bold=True, size=12, color='1F4E79')

wb = openpyxl.load_workbook(OUT_PATH)

# Remove existing LM8 sheet if present
if 'LM8_LM90' in wb.sheetnames:
    del wb['LM8_LM90']

ws = wb.create_sheet(title='LM8_LM90')
title = (f'Table LM8: 90-Day Landmark Sensitivity — Cox PH (Home+HHA, N={len(lm_slp):,}); '
         f'all SLP groups have started therapy before clock begins at day 90')
ws.append([title])
ws['A1'].font = TITLE_FONT
ws.append([])

header_row = ws.max_row + 1
for col_idx, col_name in enumerate(df_out.columns, 1):
    cell = ws.cell(row=header_row, column=col_idx, value=col_name)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

for r_idx, row in enumerate(df_out.itertuples(index=False), header_row + 1):
    fill = ALT_FILL if r_idx % 2 == 0 else None
    for c_idx, val in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.alignment = Alignment(horizontal='left')
        if fill:
            cell.fill = fill

for col in ws.columns:
    max_len = max((len(str(c.value)) if c.value else 0) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

wb.save(OUT_PATH)
print(f"Saved LM8_LM90 to {OUT_PATH}")
print(f"Sheets: {', '.join(s.title for s in wb.worksheets)}")
print(f"\nN={len(lm_slp):,}  (vs 113,172 at 30d landmark — {113172-len(lm_slp):,} additional deaths day 30-90)")
print(df_out.to_string(index=False))
