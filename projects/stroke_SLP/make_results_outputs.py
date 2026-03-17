"""
make_results_outputs.py

Generates:
  1. forest_plot.png  — primary outcomes (Aspiration PNA + PEG/G-tube), 600 DPI
  2. results_table.xlsx — all outcomes except recurrent stroke
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
from lifelines import CoxPHFitter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DB_PATH      = r"F:\CMS\cms_data.duckdb"
FOREST_PATH  = r"F:\CMS\projects\stroke_SLP\forest_plot.png"
RESULTS_PATH = r"F:\CMS\projects\stroke_SLP\results_table.xlsx"
LANDMARK     = 90
MAX_FOLLOW_LM = 1795   # 1885 - 90

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        p.DSYSRTKY, p.age_at_adm, p.sex, p.stroke_type,
        p.index_los, p.mech_vent, p.peg_placed, p.trach_placed,
        p.dschg_status, p.dysphagia_poa, p.aspiration_poa,
        p.prior_stroke, p.dementia, p.van_walraven_score,
        p.slp_outpt_any_90d, p.days_to_slp_outpt,
        o.days_to_death, o.days_to_aspiration,
        o.days_to_gtube, o.pre_stroke_tube
    FROM stroke_propensity p
    JOIN stroke_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
""").df()
con.close()
print(f"  Loaded {len(df):,} rows")

# ── Landmark + home/HHA restriction ───────────────────────────────────────────
lm = df[df['days_to_death'].isna() | (df['days_to_death'] > LANDMARK)].copy()
lm_slp = lm[lm['slp_outpt_any_90d'] == 1].copy()

def slp_group(d):
    if pd.isna(d): return None
    if d <= 14:    return 'SLP 0-14d'
    if d <= 30:    return 'SLP 15-30d'
    return 'SLP 31-90d'

lm_slp['timing_group'] = lm_slp['days_to_slp_outpt'].apply(slp_group)

def map_dschg(code):
    if pd.isna(code): return 'home'
    c = str(code).strip()
    if c in ('01', '08'): return 'home'
    if c == '06':         return 'hha'
    if c in ('03', '61'): return 'snf'
    if c == '62':         return 'irf'
    return 'other'

lm_slp['dschg_group'] = lm_slp['dschg_status'].apply(map_dschg)
lm_slp = lm_slp[lm_slp['dschg_group'].isin(['home', 'hha'])].copy()
print(f"  After home/HHA filter: {len(lm_slp):,}")

# ── Landmark-shifted outcome times ────────────────────────────────────────────
for col in ['days_to_death', 'days_to_aspiration', 'days_to_gtube']:
    lm_slp[col] = lm_slp[col].astype('float64')
    lm_slp[f'{col}_lm'] = np.where(
        lm_slp[col].notna() & (lm_slp[col] > LANDMARK),
        (lm_slp[col] - LANDMARK).clip(lower=0.5), np.nan
    )

lm_slp['censor_lm'] = np.where(
    lm_slp['days_to_death_lm'].notna(),
    lm_slp['days_to_death_lm'].clip(upper=MAX_FOLLOW_LM), MAX_FOLLOW_LM
)
lm_slp['died_lm']  = lm_slp['days_to_death_lm'].notna().astype(int)
lm_slp['asp_lm']   = lm_slp['days_to_aspiration_lm'].notna().astype(int)
lm_slp['gtube_lm'] = lm_slp['days_to_gtube_lm'].notna().astype(int)

lm_slp['days_to_aspiration_lm_filled'] = (
    lm_slp['days_to_aspiration_lm'].fillna(lm_slp['censor_lm']).clip(lower=0.5))
lm_slp['days_to_gtube_lm_filled'] = (
    lm_slp['days_to_gtube_lm'].fillna(lm_slp['censor_lm']).clip(lower=0.5))

# ── Covariates ────────────────────────────────────────────────────────────────
lm_slp['slp_15_30d'] = (lm_slp['timing_group'] == 'SLP 15-30d').astype(int)
lm_slp['slp_31_90d'] = (lm_slp['timing_group'] == 'SLP 31-90d').astype(int)

dschg_d  = pd.get_dummies(lm_slp['dschg_group'], prefix='dschg', drop_first=True)
sex_d    = pd.get_dummies(lm_slp['sex'],          prefix='sex',   drop_first=True)
stroke_d = pd.get_dummies(lm_slp['stroke_type'],  prefix='stroke',drop_first=True)
lm_slp   = pd.concat([lm_slp, dschg_d, sex_d, stroke_d], axis=1)

_cands = (
    ['slp_15_30d', 'slp_31_90d',
     'age_at_adm', 'van_walraven_score', 'index_los',
     'mech_vent', 'trach_placed', 'prior_stroke', 'dementia',
     'dysphagia_poa', 'aspiration_poa']
    + list(dschg_d.columns) + list(sex_d.columns) + list(stroke_d.columns)
)
COVS = [c for c in _cands if c in lm_slp.columns and lm_slp[c].std() > 0.05]

# ── Run Cox models ─────────────────────────────────────────────────────────────
MODELS = [
    ('Aspiration PNA',  'asp_lm',   'days_to_aspiration_lm_filled', False, 0.0),
    ('PEG/G-tube',      'gtube_lm', 'days_to_gtube_lm_filled',      True,  0.1),
    ('Mortality',       'died_lm',  'censor_lm',                    False, 0.0),
]

results = []
for label, ev_col, dur_col, excl_peg, pen in MODELS:
    sub = lm_slp.copy()
    if excl_peg:
        sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
    cov_use = list(COVS)
    if not excl_peg and 'peg_placed' not in cov_use:
        cov_use = ['peg_placed'] + cov_use

    cox_df = sub[[dur_col, ev_col] + cov_use].rename(
        columns={dur_col: 'duration', ev_col: 'event'}
    ).dropna()
    n_ev = int(cox_df['event'].sum())
    n    = len(cox_df)

    try:
        cph = CoxPHFitter(penalizer=pen)
        cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
        for tv, comp in [('slp_15_30d', '15-30d vs 0-14d'),
                         ('slp_31_90d', '31-90d vs 0-14d')]:
            if tv not in cph.summary.index: continue
            r    = cph.summary.loc[tv]
            hr   = np.exp(r['coef'])
            lo95 = np.exp(r['coef lower 95%'])
            hi95 = np.exp(r['coef upper 95%'])
            p    = r['p']
            ev_val = hr + np.sqrt(hr * abs(hr - 1))
            results.append({
                'Outcome':    label,
                'Comparison': comp,
                'N':          n,
                'Events':     n_ev,
                'Event_pct':  round(100 * n_ev / n, 1),
                'HR':         round(hr,   2),
                'CI_low':     round(lo95, 2),
                'CI_high':    round(hi95, 2),
                'p_raw':      p,
                'p_str':      '<0.0001' if p < 0.0001 else f'{p:.4f}',
                'E_value':    round(ev_val, 2),
                'excl_peg':   excl_peg,
            })
        print(f"  {label}: N={n:,}, events={n_ev:,}")
    except Exception as e:
        print(f"  ERROR {label}: {e}")

df_res = pd.DataFrame(results)


# ═══════════════════════════════════════════════════════════════════════════════
# FOREST PLOT — primary outcomes only (Aspiration PNA + PEG/G-tube)
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding forest plot...")

PRIMARY  = ['Aspiration PNA', 'PEG/G-tube']
COMPS    = ['15-30d vs 0-14d', '31-90d vs 0-14d']
C_BLUE   = '#2E5090'
C_ORANGE = '#C55A11'
COMP_COL = {'15-30d vs 0-14d': C_BLUE, '31-90d vs 0-14d': C_ORANGE}
COMP_MKR = {'15-30d vs 0-14d': 'o',    '31-90d vs 0-14d': 's'}
COMP_LBL = {'15-30d vs 0-14d': 'SLP 15–30d vs 0–14d (ref)',
            '31-90d vs 0-14d': 'SLP 31–90d vs 0–14d (ref)'}

# ── Y positions ───────────────────────────────────────────────────────────────
# Row order top-to-bottom: header, asp 15-30, asp 31-90, [gap], hdr, gtube 15-30, gtube 31-90
ROW_Y = {
    'asp_hdr':    8.2,
    ('Aspiration PNA', '15-30d vs 0-14d'): 7.3,
    ('Aspiration PNA', '31-90d vs 0-14d'): 6.4,
    'gtube_hdr':  5.1,
    ('PEG/G-tube', '15-30d vs 0-14d'):    4.2,
    ('PEG/G-tube', '31-90d vs 0-14d'):    3.3,
}
YLIM = (2.5, 9.2)

# ── Figure layout ─────────────────────────────────────────────────────────────
fig = plt.figure(figsize=(14, 6), facecolor='white')
gs  = gridspec.GridSpec(1, 3, width_ratios=[3.2, 5.0, 3.5],
                         wspace=0.0, left=0.01, right=0.99,
                         top=0.88, bottom=0.12)

ax_L = fig.add_subplot(gs[0])   # left: row labels
ax_M = fig.add_subplot(gs[1])   # middle: CI bars
ax_R = fig.add_subplot(gs[2])   # right: HR text

for ax in [ax_L, ax_M, ax_R]:
    ax.set_ylim(*YLIM)
    ax.set_yticks([])
    ax.set_facecolor('white')

for ax in [ax_L, ax_R]:
    ax.set_xticks([])
    for sp in ax.spines.values():
        sp.set_visible(False)

# ── Shading ───────────────────────────────────────────────────────────────────
for ax in [ax_L, ax_M, ax_R]:
    ax.axhspan(5.55, 8.75, color='#EBF3FB', alpha=0.6, zorder=0)  # Asp PNA
    ax.axhspan(2.75, 4.65, color='#FEF4EE', alpha=0.6, zorder=0)  # G-tube

# ── Column headers ────────────────────────────────────────────────────────────
HDR_KW = dict(fontsize=9, fontweight='bold', color='#555555', va='bottom')
ax_L.text(0.97, YLIM[1]+0.02, 'Outcome / Comparison',
          ha='right', transform=ax_L.get_yaxis_transform(), **HDR_KW)
ax_M.axvline(1.0, color='#888888', lw=1.2, linestyle='--', zorder=1)
ax_R.text(0.03, YLIM[1]+0.02, 'HR [95% CI]              p-value',
          ha='left', transform=ax_R.get_yaxis_transform(), **HDR_KW)

# ── Section headers ───────────────────────────────────────────────────────────
def sec_header(y, text):
    for ax in [ax_L, ax_M, ax_R]:
        ax.axhline(y - 0.3, color='#CCCCCC', lw=0.8, zorder=0)
    ax_L.text(0.97, y, text, ha='right', va='center', fontsize=10.5,
              fontweight='bold', color='#1F4E79',
              transform=ax_L.get_yaxis_transform())

sec_header(ROW_Y['asp_hdr'],   'Aspiration Pneumonia  (PRIMARY)')
sec_header(ROW_Y['gtube_hdr'], 'PEG / G-tube Placement  (PRIMARY)')

# ── Plot rows ─────────────────────────────────────────────────────────────────
primary_df = df_res[df_res['Outcome'].isin(PRIMARY)].copy()

for _, row in primary_df.iterrows():
    key  = (row['Outcome'], row['Comparison'])
    if key not in ROW_Y: continue
    ypos = ROW_Y[key]
    hr, lo, hi = row['HR'], row['CI_low'], row['CI_high']
    col = COMP_COL[row['Comparison']]
    mkr = COMP_MKR[row['Comparison']]

    # Left label
    comp_lbl = row['Comparison'].replace('vs', 'vs.')
    ax_L.text(0.97, ypos, f"  {comp_lbl}", ha='right', va='center',
              fontsize=9.5, color='#333333',
              transform=ax_L.get_yaxis_transform())

    # CI bar + dot on middle axes (log scale)
    ax_M.plot([lo, hi], [ypos, ypos], color=col, lw=2.0, solid_capstyle='round', zorder=3)
    ax_M.plot(hr, ypos, marker=mkr, color=col, markersize=9,
              markeredgecolor='white', markeredgewidth=0.8, zorder=4)

    # Right text
    p_str = row['p_str']
    p_fmt = p_str if p_str.startswith('<') else f'= {p_str}'
    ax_R.text(0.03, ypos,
              f"  {hr:.2f}  [{lo:.2f}\u2013{hi:.2f}]     p {p_fmt}",
              ha='left', va='center', fontsize=9, color='#1F4E79',
              transform=ax_R.get_yaxis_transform())

# ── Middle axes formatting ────────────────────────────────────────────────────
ax_M.set_xscale('log')
ax_M.set_xlim(0.55, 3.2)
ax_M.set_xticks([0.6, 0.8, 1.0, 1.25, 1.5, 2.0, 2.5, 3.0])
ax_M.set_xticklabels(['0.60', '0.80', '1.00', '1.25', '1.50',
                       '2.00', '2.50', '3.00'], fontsize=8.5)
ax_M.set_xlabel('Hazard Ratio (log scale)\nReference group: SLP 0–14 days',
                fontsize=9.5, color='#333333')
for sp in ['top', 'left', 'right']:
    ax_M.spines[sp].set_visible(False)
ax_M.spines['bottom'].set_color('#AAAAAA')
ax_M.tick_params(axis='x', length=3, color='#AAAAAA')

# ── Legend ────────────────────────────────────────────────────────────────────
from matplotlib.lines import Line2D
legend_elements = [
    Line2D([0], [0], marker='o', color='w', markerfacecolor=C_BLUE,
           markersize=9, markeredgecolor='white',
           label=COMP_LBL['15-30d vs 0-14d']),
    Line2D([0], [0], marker='s', color='w', markerfacecolor=C_ORANGE,
           markersize=9, markeredgecolor='white',
           label=COMP_LBL['31-90d vs 0-14d']),
]
ax_M.legend(handles=legend_elements, loc='upper right', fontsize=9,
            frameon=True, framealpha=0.9, edgecolor='#CCCCCC')

# ── Title ─────────────────────────────────────────────────────────────────────
fig.suptitle(
    'Adjusted Hazard Ratios: Timing of Post-Discharge SLP and Stroke Outcomes\n'
    'Home / HHA-discharged Medicare stroke patients  •  90-day landmark  •  N = 108,695',
    fontsize=11, fontweight='bold', color='#1F4E79', y=0.99, va='top'
)

plt.savefig(FOREST_PATH, dpi=600, bbox_inches='tight', facecolor='white')
plt.close()
print(f"Saved: {FOREST_PATH}")


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL RESULTS TABLE — all outcomes (no recurrent stroke)
# ═══════════════════════════════════════════════════════════════════════════════
print("\nBuilding results table Excel...")

HEADER_FILL   = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT   = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL  = PatternFill('solid', fgColor='BDD7EE')
SECTION_FONT  = Font(bold=True, size=10, color='1F4E79')
ALT_FILL      = PatternFill('solid', fgColor='EBF3FB')
TITLE_FONT    = Font(bold=True, size=12, color='1F4E79')
BOLD_FONT     = Font(bold=True, size=10)
SIG_FONT      = Font(bold=True, size=10, color='C55A11')   # highlight significant
BORDER_BOT    = Border(bottom=Side(style='thin', color='CCCCCC'))

OUTCOME_LABELS = {
    'Aspiration PNA': 'Aspiration Pneumonia',
    'PEG/G-tube':     'PEG / G-tube Placement',
    'Mortality':      'All-cause Mortality',
}
OUTCOME_NOTES = {
    'Aspiration PNA': 'ICD-10 J690/J698; time-to-first event from 90-day landmark',
    'PEG/G-tube':     'Excludes patients with index PEG or pre-stroke tube; penalized Cox (pen=0.1)',
    'Mortality':      'Death from MBSF DEATH_DT; max follow-up 1,795 days from landmark',
}

# Build display DataFrame
display_rows = []
for outcome in ['Aspiration PNA', 'PEG/G-tube', 'Mortality']:
    # Section header
    display_rows.append({
        '_section': True,
        'Outcome': OUTCOME_LABELS[outcome],
        'Comparison': '',
        'N': '', 'Events': '', 'Event %': '',
        'HR (95% CI)': '', 'p-value': '', 'E-value': '',
        'Note': OUTCOME_NOTES[outcome],
    })
    sub = df_res[df_res['Outcome'] == outcome]
    for comp in ['15-30d vs 0-14d', '31-90d vs 0-14d']:
        row = sub[sub['Comparison'] == comp]
        if row.empty:
            continue
        r = row.iloc[0]
        display_rows.append({
            '_section': False,
            'Outcome': '',
            'Comparison': comp,
            'N': f"{r['N']:,}",
            'Events': f"{r['Events']:,}",
            'Event %': f"{r['Event_pct']:.1f}%",
            'HR (95% CI)': f"{r['HR']:.2f}  [{r['CI_low']:.2f}\u2013{r['CI_high']:.2f}]",
            'p-value': r['p_str'],
            'E-value': f"{r['E_value']:.2f}",
            'Note': '',
        })

display_cols = ['Outcome', 'Comparison', 'N', 'Events', 'Event %',
                'HR (95% CI)', 'p-value', 'E-value', 'Note']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Results'

# Title
ws.append(['Table: Adjusted Cox PH Results — SLP Timing and Post-Stroke Outcomes'])
ws['A1'].font = TITLE_FONT
ws.append(['Home / HHA-discharged Medicare stroke patients  •  90-day landmark  •  Reference: SLP 0–14 days'])
ws['A2'].font = Font(italic=True, size=10, color='555555')
ws.append([])

# Column headers
hdr_row = ws.max_row + 1
for ci, col in enumerate(display_cols, 1):
    cell = ws.cell(row=hdr_row, column=ci, value=col)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
ws.row_dimensions[hdr_row].height = 28

# Data rows
alt = 0
for dr in display_rows:
    is_sec = dr['_section']
    ri = ws.max_row + 1
    if not is_sec:
        alt += 1
    for ci, col in enumerate(display_cols, 1):
        val  = dr.get(col, '')
        cell = ws.cell(row=ri, column=ci, value=val)
        cell.alignment = Alignment(
            horizontal='center' if ci > 1 else 'left',
            vertical='center', wrap_text=True
        )
        if is_sec:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif alt % 2 == 0:
            cell.fill = ALT_FILL

        # Bold + orange for significant p-values
        if col == 'p-value' and not is_sec and val == '<0.0001':
            cell.font = SIG_FONT
        elif col == 'HR (95% CI)' and not is_sec:
            cell.font = BOLD_FONT

    ws.row_dimensions[ri].height = 18

# Footnote
fn_row = ws.max_row + 2
ws.cell(row=fn_row, column=1,
        value=('Covariates: age, van Walraven score, LOS, mechanical ventilation, '
               'tracheostomy, prior stroke, dementia, dysphagia POA, aspiration POA, '
               'discharge destination, sex, stroke type.  '
               'E-value: minimum unmeasured confounding needed to explain away result '
               '(VanderWeele & Ding 2017).'))
ws.cell(row=fn_row, column=1).font = Font(italic=True, size=8.5, color='666666')
ws.merge_cells(start_row=fn_row, start_column=1,
               end_row=fn_row,   end_column=len(display_cols))

# Column widths
col_widths = [28, 20, 10, 10, 10, 24, 10, 10, 40]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

ws.freeze_panes = f'A{hdr_row + 1}'
wb.save(RESULTS_PATH)
print(f"Saved: {RESULTS_PATH}")
