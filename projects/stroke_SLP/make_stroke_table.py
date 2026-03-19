"""
make_stroke_table.py

Exports stroke + SLP timing analysis results to a formatted Excel workbook:
  F:\CMS\projects\stroke_SLP\stroke_slp_tables.xlsx

Two pairwise comparisons:
  Comparison A: 0-14d  vs 31-90d
  Comparison B: 15-30d vs 31-90d

Sheets:
  Table1_Cohort        — Cohort characteristics by timing group (unmatched)
  Table2_Balance_A     — Covariate balance comparison A (pre/post SMDs)
  Table2_Balance_B     — Covariate balance comparison B (pre/post SMDs)
  Table3_OR_A          — OR table: all outcomes, matched comparison A
  Table3_OR_B          — OR table: all outcomes, matched comparison B
  Table4_ByStrokeType_A / _B  — OR stratified by stroke type
  Table5_ByAge_A / _B         — OR stratified by age group
  Table5b_ByDischg_A / _B     — OR stratified by discharge disposition (Home / Home+HHA / SNF / IRF / LTACH)
  Table6_CoxPH_A / _B        — Cox HR (mortality, aspiration, dysphagia)
  Table7_Costs_A / _B        — Medicare cost analysis
  Table8_KM_A / _B           — KM survival / cumulative incidence
"""

import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
from scipy.stats import chi2_contingency, mannwhitneyu
from lifelines import CoxPHFitter, KaplanMeierFitter
from lifelines.statistics import logrank_test
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"F:\CMS\projects\stroke_SLP\stroke_slp_tables.xlsx"

MAX_FOLLOW = 365
TIMEPOINTS = [90, 180, 365]

COMPARISONS = [
    ('A', '0-14d',  '31-90d', 'psm_matched_A'),
    ('B', '15-30d', '31-90d', 'psm_matched_B'),
]

TIMING_ORDER = ['0-14d', '15-30d', '31-90d', 'No SLP']

# ── Style helpers ──────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
BOLD        = Font(bold=True)
WHITE_BOLD  = Font(bold=True, color="FFFFFF")
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin",   color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, sub=False):
    cell.fill      = SUBHDR_FILL if sub else HEADER_FILL
    cell.font      = WHITE_BOLD
    cell.alignment = CENTER
    cell.border    = THIN_BORDER


def style_data(cell, alt=False):
    cell.fill      = ALT_FILL if alt else PatternFill()
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def autofit(ws, min_width=8, max_width=40):
    for col in ws.columns:
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, min_width), max_width)


# ── Load full cohort (for Table 1 and SMD pre-match) ─────────────────────────

def load_full(con):
    return con.execute("""
        SELECT
            p.DSYSRTKY,
            p.slp_timing_group,
            p.psm_matched_A,
            p.psm_matched_B,
            p.age_at_adm,
            p.sex,
            p.race,
            p.stroke_type,
            p.adm_year,
            p.index_los,
            p.dysphagia_poa,
            p.aspiration_poa,
            p.mech_vent,
            p.peg_placed,
            p.trach_placed,
            p.prior_stroke,
            p.van_walraven_score,
            p.afib,
            p.hypertension,
            c.index_pmt
        FROM stroke_propensity p
        JOIN stroke_cohort c ON c.DSYSRTKY = p.DSYSRTKY
    """).df()


def load_matched(con, match_col):
    return con.execute(f"""
        SELECT
            p.DSYSRTKY,
            p.slp_timing_group,
            p.dschg_group,
            p.age_at_adm,
            p.sex,
            p.race,
            p.stroke_type,
            p.adm_year,
            p.index_los,
            p.dysphagia_poa,
            p.aspiration_poa,
            p.mech_vent,
            p.peg_placed,
            p.trach_placed,
            p.prior_stroke,
            p.van_walraven_score,
            p.afib,
            p.hypertension,
            o.days_to_death,
            o.days_to_readmit,
            o.n_readmissions_365d,
            o.days_to_recur_stroke,
            o.days_to_aspiration,
            o.days_to_dysphagia,
            o.days_to_gtube,
            o.days_to_snf,
            o.snf_30d,
            o.hha_90d,
            o.total_pmt_365d,
            o.snf_pmt_365d,
            o.hha_pmt_365d,
            c.index_pmt
        FROM stroke_propensity p
        JOIN stroke_outcomes   o ON o.DSYSRTKY = p.DSYSRTKY
        JOIN stroke_cohort     c ON c.DSYSRTKY = p.DSYSRTKY
        WHERE p.{match_col} = TRUE
    """).df()


# ── Derived columns ────────────────────────────────────────────────────────────

def add_derived(df, treat_grp):
    df = df.copy()
    df['treated'] = (df['slp_timing_group'] == treat_grp).astype(int)
    df['censor_days'] = np.where(
        df['days_to_death'].notna(),
        df['days_to_death'].clip(upper=MAX_FOLLOW),
        MAX_FOLLOW
    )
    for days, lbl in [(90,'90d'),(180,'180d'),(365,'365d')]:
        df[f'died_{lbl}']      = (df['days_to_death'].notna()        & (df['days_to_death']        <= days)).astype(int)
        df[f'readmit_{lbl}']   = (df['days_to_readmit'].notna()      & (df['days_to_readmit']      <= days)).astype(int)
        df[f'recur_{lbl}']     = (df['days_to_recur_stroke'].notna() & (df['days_to_recur_stroke'] <= days)).astype(int)
        df[f'asp_{lbl}']       = (df['days_to_aspiration'].notna()   & (df['days_to_aspiration']   <= days)).astype(int)
        df[f'dysphagia_{lbl}'] = (df['days_to_dysphagia'].notna()    & (df['days_to_dysphagia']    <= days)).astype(int)
        df[f'gtube_{lbl}']     = (df['days_to_gtube'].notna()        & (df['days_to_gtube']        <= days)).astype(int)
    return df


# ── OR computation ─────────────────────────────────────────────────────────────

def compute_or(sub, ev_col, days_col, cutoff):
    if cutoff is None:
        elig = sub.copy()
        elig['ev'] = elig[ev_col].astype(int)
    else:
        mask = (
            (sub['censor_days'] >= cutoff) |
            (sub[ev_col].astype(bool) & (sub[days_col] <= cutoff))
        )
        elig = sub[mask].copy()
        elig['ev'] = (elig[ev_col].astype(bool) & (elig[days_col] <= cutoff)).astype(int)

    t_ = elig[elig['treated'] == 1]
    c_ = elig[elig['treated'] == 0]
    nt, nc = len(t_), len(c_)
    et, ec = int(t_['ev'].sum()), int(c_['ev'].sum())

    a, b, c, d = et, nt - et, ec, nc - ec
    if 0 in (a, b, c, d) or nt < 10 or nc < 10:
        return dict(n_treat=nt, ev_treat=et, pct_treat="N/A",
                    n_ctrl=nc, ev_ctrl=ec, pct_ctrl="N/A",
                    OR="N/A", CI="N/A", p="N/A")

    or_v   = (a * d) / (b * c)
    log_or = np.log(or_v)
    se     = np.sqrt(1/a + 1/b + 1/c + 1/d)
    lo, hi = np.exp(log_or - 1.96*se), np.exp(log_or + 1.96*se)
    _, p, _, _ = chi2_contingency([[a, b], [c, d]], correction=False)
    p_str = "<0.001" if p < 0.001 else f"{p:.3f}"
    return dict(
        n_treat=nt, ev_treat=et, pct_treat=f"{100*et/nt:.1f}%",
        n_ctrl=nc, ev_ctrl=ec, pct_ctrl=f"{100*ec/nc:.1f}%",
        OR=f"{or_v:.2f}", CI=f"[{lo:.2f}, {hi:.2f}]", p=p_str
    )


def build_or_rows(sub, strat_label):
    outcomes = [
        ("Mortality",        "died_{tp}",      "days_to_death",        TIMEPOINTS),
        ("Readmission",      "readmit_{tp}",   "days_to_readmit",      TIMEPOINTS),
        ("Aspiration PNA",   "asp_{tp}",       "days_to_aspiration",   TIMEPOINTS),
        ("Dysphagia Dx",     "dysphagia_{tp}", "days_to_dysphagia",    TIMEPOINTS),
        ("G-tube",           "gtube_{tp}",     "days_to_gtube",        TIMEPOINTS),
        ("SNF 30d",          "snf_30d",        "days_to_snf",          [None]),
        ("HHA 90d",          "hha_90d",        "days_to_snf",          [None]),
    ]
    rows = []
    for lbl_o, ev_tmpl, days_col, cutoffs in outcomes:
        for cutoff in cutoffs:
            if '{tp}' in ev_tmpl:
                tp_lbl = f"{cutoff}d"
                ev_col = ev_tmpl.replace('{tp}', tp_lbl)
                display = f"{lbl_o} {tp_lbl}"
            else:
                ev_col = ev_tmpl
                display = lbl_o
            d = compute_or(sub, ev_col, days_col, cutoff)
            rows.append({"Stratum": strat_label, "Outcome": display, **d})
    return rows


def write_or_sheet(wb, sheet_name, rows, treat_grp, ctrl_grp):
    ws = wb.create_sheet(sheet_name)
    col_display = ["Stratum", "Outcome",
                   f"n ({treat_grp})", f"Events ({treat_grp})", f"% ({treat_grp})",
                   f"n ({ctrl_grp})", f"Events ({ctrl_grp})", f"% ({ctrl_grp})",
                   "OR", "95% CI", "p"]
    col_keys = ["Stratum", "Outcome",
                "n_treat", "ev_treat", "pct_treat",
                "n_ctrl", "ev_ctrl", "pct_ctrl",
                "OR", "CI", "p"]
    for ci, h in enumerate(col_display, 1):
        style_header(ws.cell(row=1, column=ci, value=h))
    for ri, row in enumerate(rows, 2):
        alt = ri % 2 == 0
        for ci, key in enumerate(col_keys, 1):
            style_data(ws.cell(row=ri, column=ci, value=row.get(key, "")), alt)
    autofit(ws)
    ws.freeze_panes = "A2"


# ── SMD helper ────────────────────────────────────────────────────────────────

def _smd(x0, x1):
    mu0, mu1 = x0.mean(), x1.mean()
    s0,  s1  = x0.std(),  x1.std()
    pooled   = np.sqrt((s0**2 + s1**2) / 2)
    return abs(mu1 - mu0) / pooled if pooled > 0 else 0.0


BALANCE_VARS = [
    ('age_at_adm',        'Age at admission'),
    ('index_los',         'Index LOS'),
    ('van_walraven_score','van Walraven score'),
    ('dysphagia_poa',     'Dysphagia POA'),
    ('aspiration_poa',    'Aspiration POA'),
    ('mech_vent',         'Mechanical ventilation'),
    ('peg_placed',        'PEG placed'),
    ('trach_placed',      'Tracheostomy'),
    ('prior_stroke',      'Prior stroke'),
    ('afib',              'Atrial fibrillation'),
    ('hypertension',      'Hypertension'),
]


def write_balance_sheet(wb, sheet_name, full_df, matched_df, treat_grp, ctrl_grp):
    ws = wb.create_sheet(sheet_name)
    hdr = ["Variable", "Pre-match SMD", "Post-match SMD", "Balance OK (<0.1)"]
    for ci, h in enumerate(hdr, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    pre_t  = full_df[full_df['slp_timing_group'] == treat_grp]
    pre_c  = full_df[full_df['slp_timing_group'] == ctrl_grp]
    post_t = matched_df[matched_df['treated'] == 1]
    post_c = matched_df[matched_df['treated'] == 0]

    for ri, (col, label) in enumerate(BALANCE_VARS, 2):
        if col not in full_df.columns:
            continue
        smd_pre  = _smd(pre_c[col].fillna(0),  pre_t[col].fillna(0))
        smd_post = _smd(post_c[col].fillna(0), post_t[col].fillna(0))
        ok = "Yes" if smd_post < 0.1 else "No  ***"
        alt = ri % 2 == 0
        for ci, val in enumerate([label, round(smd_pre, 3), round(smd_post, 3), ok], 1):
            style_data(ws.cell(row=ri, column=ci, value=val), alt)

    autofit(ws)
    ws.freeze_panes = "A2"


# ── Table 1: Cohort characteristics ───────────────────────────────────────────

def write_table1(wb, full_df):
    ws = wb.create_sheet("Table1_Cohort")

    groups = [g for g in TIMING_ORDER if g in full_df['slp_timing_group'].values]
    hdr = ["Variable"] + groups
    for ci, h in enumerate(hdr, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    subs = {g: full_df[full_df['slp_timing_group'] == g] for g in groups}

    def _pct(sub, col):
        n = len(sub)
        s = sub[col].fillna(0).sum()
        return f"{int(s):,} ({100*s/n:.1f}%)" if n > 0 else "N/A"

    def _mean_sd(sub, col):
        v = sub[col].dropna()
        return f"{v.mean():.1f} ± {v.std():.1f}" if len(v) > 0 else "N/A"

    def _cat(sub, col, val):
        n = len(sub)
        s = (sub[col] == val).sum()
        return f"{s:,} ({100*s/n:.1f}%)" if n > 0 else "N/A"

    table_rows = [
        ("N",                        lambda s: f"{len(s):,}"),
        ("Age, mean (SD)",            lambda s: _mean_sd(s, 'age_at_adm')),
        ("Male sex, n (%)",           lambda s: _cat(s, 'sex', 'Male')),
        ("Race — White, n (%)",       lambda s: _cat(s, 'race', 'White')),
        ("Race — Black, n (%)",       lambda s: _cat(s, 'race', 'Black')),
        ("Stroke — Ischemic, n (%)",  lambda s: _cat(s, 'stroke_type', 'Ischemic')),
        ("Stroke — ICH, n (%)",       lambda s: _cat(s, 'stroke_type', 'ICH')),
        ("Stroke — SAH, n (%)",       lambda s: _cat(s, 'stroke_type', 'SAH')),
        ("Index LOS, mean (SD)",      lambda s: _mean_sd(s, 'index_los')),
        ("van Walraven, mean (SD)",   lambda s: _mean_sd(s, 'van_walraven_score')),
        ("Dysphagia POA, n (%)",      lambda s: _pct(s, 'dysphagia_poa')),
        ("Aspiration POA, n (%)",     lambda s: _pct(s, 'aspiration_poa')),
        ("Mech. ventilation, n (%)",  lambda s: _pct(s, 'mech_vent')),
        ("PEG placed, n (%)",         lambda s: _pct(s, 'peg_placed')),
        ("Tracheostomy, n (%)",       lambda s: _pct(s, 'trach_placed')),
        ("Prior stroke, n (%)",       lambda s: _pct(s, 'prior_stroke')),
        ("Atrial fibrillation, n (%)",lambda s: _pct(s, 'afib')),
        ("Hypertension, n (%)",       lambda s: _pct(s, 'hypertension')),
    ]

    for ri, (label, fn) in enumerate(table_rows, 2):
        alt = ri % 2 == 0
        row_vals = [label] + [fn(subs[g]) for g in groups]
        for ci, v in enumerate(row_vals, 1):
            style_data(ws.cell(row=ri, column=ci, value=v), alt)

    autofit(ws)
    ws.freeze_panes = "B2"


# ── Cox PH sheet ───────────────────────────────────────────────────────────────

def write_cox_sheet(wb, sheet_name, matched_df):
    ws = wb.create_sheet(sheet_name)
    hdr = ["Outcome", "HR", "95% CI Lower", "95% CI Upper", "p", "N", "Events"]
    for ci, h in enumerate(hdr, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    cox_outcomes = [
        ("All-cause mortality (1yr)", "died_365d",       "censor_days"),
        ("Aspiration PNA (1yr)",      "asp_365d",        "days_to_aspiration"),
        ("Dysphagia Dx (1yr)",        "dysphagia_365d",  "days_to_dysphagia"),
        ("G-tube (1yr)",              "gtube_365d",      "days_to_gtube"),
    ]
    covariates = ['treated', 'age_at_adm', 'van_walraven_score', 'dysphagia_poa',
                  'aspiration_poa', 'mech_vent', 'peg_placed', 'trach_placed', 'index_los']

    for ri, (lbl_c, ev_col, dur_col) in enumerate(cox_outcomes, 2):
        sub = matched_df.copy()
        sub['_dur'] = sub.get(dur_col, sub['censor_days']).fillna(sub['censor_days']).astype(float).clip(lower=0.5)
        sub['_ev']  = sub[ev_col].astype(int)
        cox_df = sub[['_dur', '_ev'] + covariates].dropna()
        cox_df = cox_df.rename(columns={'_dur': 'duration', '_ev': 'event'})
        try:
            cph = CoxPHFitter()
            cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
            row  = cph.summary.loc['treated']
            hr   = round(float(np.exp(row['coef'])), 2)
            lo95 = round(float(np.exp(row['coef lower 95%'])), 2)
            hi95 = round(float(np.exp(row['coef upper 95%'])), 2)
            p    = round(float(row['p']), 4)
            vals = [lbl_c, hr, lo95, hi95, p, len(cox_df), int(cox_df['event'].sum())]
        except Exception as e:
            vals = [lbl_c, "ERROR", str(e), "", "", "", ""]
        alt = ri % 2 == 0
        for ci, v in enumerate(vals, 1):
            style_data(ws.cell(row=ri, column=ci, value=v), alt)

    autofit(ws)
    ws.freeze_panes = "A2"


# ── Cost sheet ─────────────────────────────────────────────────────────────────

def write_cost_sheet(wb, sheet_name, matched_df, treat_grp, ctrl_grp):
    ws = wb.create_sheet(sheet_name)
    hdr = ["Group", "N", "Mean 1yr Cost ($)", "Median 1yr Cost ($)",
           "Mean Index Cost ($)", "Median Index Cost ($)", "Mann-Whitney p"]
    for ci, h in enumerate(hdr, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    c1 = matched_df[matched_df['treated'] == 1]['total_pmt_365d'].dropna()
    c0 = matched_df[matched_df['treated'] == 0]['total_pmt_365d'].dropna()
    i1 = matched_df[matched_df['treated'] == 1]['index_pmt'].dropna()
    i0 = matched_df[matched_df['treated'] == 0]['index_pmt'].dropna()

    if len(c1) >= 10 and len(c0) >= 10:
        _, mw_p = mannwhitneyu(c1, c0, alternative='two-sided')
        p_str = "<0.001" if mw_p < 0.001 else f"{mw_p:.3f}"
    else:
        p_str = "N/A"

    for ri, (grp, c, idx_c) in enumerate([(treat_grp, c1, i1), (ctrl_grp, c0, i0)], 2):
        alt = ri % 2 == 0
        p_val = p_str if ri == 2 else ""
        row_vals = [grp, len(c),
                    round(c.mean(), 0) if len(c) > 0 else "N/A",
                    round(c.median(), 0) if len(c) > 0 else "N/A",
                    round(idx_c.mean(), 0) if len(idx_c) > 0 else "N/A",
                    round(idx_c.median(), 0) if len(idx_c) > 0 else "N/A",
                    p_val]
        for ci, v in enumerate(row_vals, 1):
            style_data(ws.cell(row=ri, column=ci, value=v), alt)

    autofit(ws)


# ── KM sheet ───────────────────────────────────────────────────────────────────

def km_at(kmf, t):
    s  = float(kmf.predict(t))
    ci = kmf.confidence_interval_survival_function_
    idx = min(ci.index.searchsorted(t, side='right'), len(ci) - 1)
    return s, float(ci.iloc[idx, 0]), float(ci.iloc[idx, 1])


def write_km_sheet(wb, sheet_name, matched_df, treat_grp, ctrl_grp):
    ws = wb.create_sheet(sheet_name)
    hdr = ["Outcome", "Metric", "Time (days)",
           f"% ({treat_grp})", f"CI Lo ({treat_grp})", f"CI Hi ({treat_grp})",
           f"% ({ctrl_grp})", f"CI Lo ({ctrl_grp})", f"CI Hi ({ctrl_grp})",
           "Log-rank p"]
    for ci, h in enumerate(hdr, 1):
        style_header(ws.cell(row=1, column=ci, value=h))

    km_outcomes = [
        ('All-cause mortality',  'died_365d',       'censor_days',         True),
        ('Aspiration PNA',       'asp_365d',        'days_to_aspiration',  False),
        ('Dysphagia Dx',         'dysphagia_365d',  'days_to_dysphagia',   False),
        ('G-tube',               'gtube_365d',      'days_to_gtube',       False),
    ]

    ri = 2
    for lbl_km, ev_col, dur_col, show_surv in km_outcomes:
        sub = matched_df.copy()
        sub['_dur'] = sub[dur_col].fillna(sub['censor_days']).astype(float).clip(lower=0.5)
        sub['_ev']  = sub[ev_col].astype(int)

        g1 = sub[sub['treated'] == 1]
        g0 = sub[sub['treated'] == 0]

        kmf1 = KaplanMeierFitter()
        kmf0 = KaplanMeierFitter()
        kmf1.fit(g1['_dur'], g1['_ev'])
        kmf0.fit(g0['_dur'], g0['_ev'])

        lr    = logrank_test(g1['_dur'], g0['_dur'],
                             event_observed_A=g1['_ev'], event_observed_B=g0['_ev'])
        p_str = "<0.001" if lr.p_value < 0.001 else f"{lr.p_value:.4f}"
        metric = "Survival %" if show_surv else "Cum. incidence %"

        for t in TIMEPOINTS:
            s1, lo1, hi1 = km_at(kmf1, t)
            s0, lo0, hi0 = km_at(kmf0, t)
            if show_surv:
                v1 = (round(100*s1, 1), round(100*lo1, 1), round(100*hi1, 1))
                v0 = (round(100*s0, 1), round(100*lo0, 1), round(100*hi0, 1))
            else:
                v1 = (round(100*(1-s1), 1), round(100*(1-hi1), 1), round(100*(1-lo1), 1))
                v0 = (round(100*(1-s0), 1), round(100*(1-hi0), 1), round(100*(1-lo0), 1))

            alt = ri % 2 == 0
            row_vals = [lbl_km, metric, t,
                        v1[0], v1[1], v1[2],
                        v0[0], v0[1], v0[2],
                        p_str if t == TIMEPOINTS[0] else ""]
            for ci_col, v in enumerate(row_vals, 1):
                style_data(ws.cell(row=ri, column=ci_col, value=v), alt)
            ri += 1

    autofit(ws)
    ws.freeze_panes = "A2"


# ── Main ───────────────────────────────────────────────────────────────────────

print("Connecting ...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")

print("Loading full cohort ...")
full_df = load_full(con)
print(f"  Full cohort: {len(full_df):,}")
print("  Timing distribution:\n  " +
      "  ".join(f"{g}:{(full_df['slp_timing_group']==g).sum():,}" for g in TIMING_ORDER))

wb = openpyxl.Workbook()
wb.remove(wb.active)

print("Building Table 1 (cohort characteristics) ...")
write_table1(wb, full_df)

for comp_label, treat_grp, ctrl_grp, match_col in COMPARISONS:
    print(f"\nLoading matched cohort ({comp_label}: {treat_grp} vs {ctrl_grp}) ...")
    matched_raw = load_matched(con, match_col)
    matched_df  = add_derived(matched_raw, treat_grp)
    n_treat = (matched_df['treated'] == 1).sum()
    n_ctrl  = (matched_df['treated'] == 0).sum()
    print(f"  n={len(matched_df):,}  treated={n_treat:,}  control={n_ctrl:,}")

    print(f"  Building balance sheet ({comp_label}) ...")
    write_balance_sheet(wb, f"Table2_Balance_{comp_label}", full_df, matched_df, treat_grp, ctrl_grp)

    print(f"  Building OR tables ({comp_label}) ...")
    rows_all = build_or_rows(matched_df, "All matched")
    write_or_sheet(wb, f"Table3_OR_{comp_label}", rows_all, treat_grp, ctrl_grp)

    rows_stype = []
    for stype in ['Ischemic', 'ICH', 'SAH', 'Unspecified']:
        sub = matched_df[matched_df['stroke_type'] == stype]
        if len(sub) > 20:
            rows_stype.extend(build_or_rows(sub, stype))
    write_or_sheet(wb, f"Table4_ByStrokeType_{comp_label}", rows_stype, treat_grp, ctrl_grp)

    rows_age = []
    for lbl, sub in [('<75',  matched_df[matched_df['age_at_adm'] < 75]),
                     ('75-84',matched_df[(matched_df['age_at_adm'] >= 75) & (matched_df['age_at_adm'] < 85)]),
                     ('85+',  matched_df[matched_df['age_at_adm'] >= 85])]:
        rows_age.extend(build_or_rows(sub, f"Age {lbl}"))
    write_or_sheet(wb, f"Table5_ByAge_{comp_label}", rows_age, treat_grp, ctrl_grp)

    # Discharge disposition strata — same-group-vs-same-group comparisons
    # (matching was exact on dschg_group so these subsets are internally balanced)
    rows_dschg = []
    for grp in ['Home', 'Home+HHA', 'SNF', 'IRF', 'LTACH']:
        sub = matched_df[matched_df['dschg_group'] == grp]
        if len(sub) > 20:
            rows_dschg.extend(build_or_rows(sub, grp))
    write_or_sheet(wb, f"Table5b_ByDischg_{comp_label}", rows_dschg, treat_grp, ctrl_grp)

    print(f"  Building Cox PH sheet ({comp_label}) ...")
    write_cox_sheet(wb, f"Table6_CoxPH_{comp_label}", matched_df)

    print(f"  Building cost sheet ({comp_label}) ...")
    write_cost_sheet(wb, f"Table7_Costs_{comp_label}", matched_df, treat_grp, ctrl_grp)

    print(f"  Building KM sheet ({comp_label}) ...")
    write_km_sheet(wb, f"Table8_KM_{comp_label}", matched_df, treat_grp, ctrl_grp)

con.close()

wb.save(OUT_PATH)
print(f"\nSaved: {OUT_PATH}")
sheets = [ws.title for ws in wb.worksheets]
print("Sheets: " + ", ".join(sheets))
