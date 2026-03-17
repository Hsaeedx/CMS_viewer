"""
make_table1.py
Cohort characteristics table for presentation.
Columns: Full stroke cohort | Study cohort (home/HHA, SLP, survived 90d) | 0-14d | 15-30d | 31-90d
Output: F:\CMS\projects\stroke_SLP\table1_psm.xlsx
"""
import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb, numpy as np, pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"F:\CMS\projects\stroke_SLP\table1_psm.xlsx"

print("Loading data...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        p.DSYSRTKY,
        p.dschg_status,
        p.slp_outpt_any_90d,
        p.slp_outpt_0_14d, p.slp_outpt_15_30d, p.slp_outpt_31_90d,
        p.age_at_adm, p.sex, p.stroke_type,
        p.index_los, p.van_walraven_score,
        p.dysphagia_poa, p.aspiration_poa,
        p.mech_vent, p.peg_placed, p.trach_placed,
        p.prior_stroke, p.dementia, p.afib, p.hypertension,
        o.days_to_death
    FROM stroke_propensity p
    JOIN stroke_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
""").df()
con.close()
print(f"  Loaded {len(df):,} rows")

def map_dschg(c):
    c = str(c).strip() if not pd.isna(c) else ''
    if c in ('01','08'): return 'home'
    if c == '06':        return 'hha'
    if c in ('03','61'): return 'snf'
    if c == '62':        return 'irf'
    return 'other'

df['dschg_group'] = df['dschg_status'].apply(map_dschg)
df['days_to_death'] = df['days_to_death'].astype('float64')

def slp_timing(row):
    if row['slp_outpt_0_14d']  == 1: return '0-14d'
    if row['slp_outpt_15_30d'] == 1: return '15-30d'
    if row['slp_outpt_31_90d'] == 1: return '31-90d'
    return None

df['timing'] = df.apply(slp_timing, axis=1)

# ── Define cohorts ─────────────────────────────────────────────────────────────
full    = df.copy()
study   = df[
    (df['slp_outpt_any_90d'] == 1) &
    (df['days_to_death'].isna() | (df['days_to_death'] > 90)) &
    (df['dschg_group'].isin(['home','hha']))
].copy()
g0 = study[study['timing'] == '0-14d']
g1 = study[study['timing'] == '15-30d']
g2 = study[study['timing'] == '31-90d']

cohorts = [
    ('Full Cohort',   full),
    ('Study Cohort',  study),
    ('SLP 0-14d',     g0),
    ('SLP 15-30d',    g1),
    ('SLP 31-90d',    g2),
]
print(f"  Full={len(full):,}  Study={len(study):,}  "
      f"0-14d={len(g0):,}  15-30d={len(g1):,}  31-90d={len(g2):,}")

# ── Helper ─────────────────────────────────────────────────────────────────────
def fmt_mean_sd(s):   return f"{s.mean():.1f} ({s.std():.1f})"
def fmt_pct(s, v=None):
    p = 100.0*(s==v).mean() if v is not None else 100.0*s.mean()
    return f"{p:.1f}%"
def fmt_n(s):         return f"{len(s):,}"

# ── Build rows ─────────────────────────────────────────────────────────────────
rows = []

def section(title):
    rows.append([title] + ['']*len(cohorts))

def row(label, fn, indent=False):
    r = [('  ' + label) if indent else label]
    for _, c in cohorts:
        try:    r.append(fn(c))
        except: r.append('—')
    rows.append(r)

headers = ['Characteristic'] + [f"{name}\n(N={fmt_n(c)})" for name, c in cohorts]

section('DEMOGRAPHICS')
row('Age, mean (SD)',             lambda c: fmt_mean_sd(c['age_at_adm']))
row('Female, %',                  lambda c: fmt_pct(c['sex'], 'Female'))

section('STROKE TYPE')
row('Ischemic, %',                lambda c: fmt_pct(c['stroke_type'], 'Ischemic'), indent=True)
row('Intracerebral hemorrhage, %',lambda c: fmt_pct(c['stroke_type'], 'ICH'),      indent=True)
row('Subarachnoid hemorrhage, %', lambda c: fmt_pct(c['stroke_type'], 'SAH'),      indent=True)

section('HOSPITAL COURSE')
row('Index LOS, mean (SD) days',  lambda c: fmt_mean_sd(c['index_los']))
row('Mechanical ventilation, %',  lambda c: fmt_pct(c['mech_vent']))
row('Tracheostomy, %',            lambda c: fmt_pct(c['trach_placed']))
row('PEG at admission, %',        lambda c: fmt_pct(c['peg_placed']))

section('SWALLOWING / SPEECH')
row('Dysphagia present on admission, %', lambda c: fmt_pct(c['dysphagia_poa']))
row('Aspiration present on admission, %',lambda c: fmt_pct(c['aspiration_poa']))

section('COMORBIDITIES')
row('van Walraven score, mean (SD)', lambda c: fmt_mean_sd(c['van_walraven_score']))
row('Prior stroke, %',             lambda c: fmt_pct(c['prior_stroke']))
row('Dementia, %',                 lambda c: fmt_pct(c['dementia']))
row('Atrial fibrillation, %',      lambda c: fmt_pct(c['afib']))
row('Hypertension, %',             lambda c: fmt_pct(c['hypertension']))

section('DISCHARGE DESTINATION')
row('Home, %',                    lambda c: fmt_pct(c['dschg_group'], 'home'))
row('Home health agency, %',      lambda c: fmt_pct(c['dschg_group'], 'hha'))
row('SNF, %',                     lambda c: fmt_pct(c['dschg_group'], 'snf'))
row('IRF, %',                     lambda c: fmt_pct(c['dschg_group'], 'irf'))

df_table = pd.DataFrame(rows, columns=headers)

# ── Write Excel ────────────────────────────────────────────────────────────────
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
SECTION_FILL = PatternFill('solid', fgColor='BDD7EE')
SECTION_FONT = Font(bold=True, size=10, color='1F4E79')
ALT_FILL     = PatternFill('solid', fgColor='EBF3FB')
TITLE_FONT   = Font(bold=True, size=12, color='1F4E79')

SECTION_LABELS = {
    'DEMOGRAPHICS','STROKE TYPE','HOSPITAL COURSE',
    'SWALLOWING / SPEECH','COMORBIDITIES','DISCHARGE DESTINATION'
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Table1'

ws.append(['Table 1: Cohort Characteristics'])
ws['A1'].font = TITLE_FONT
ws.append([])

header_row = ws.max_row + 1
for ci, cn in enumerate(df_table.columns, 1):
    cell = ws.cell(row=header_row, column=ci, value=cn)
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

alt = 0
for ri, row_data in enumerate(df_table.itertuples(index=False), header_row + 1):
    char = row_data[0]
    is_sec = char.strip() in SECTION_LABELS
    if not is_sec: alt += 1
    for ci, val in enumerate(row_data, 1):
        cell = ws.cell(row=ri, column=ci, value=val)
        if is_sec:
            cell.font = SECTION_FONT
            cell.fill = SECTION_FILL
        elif alt % 2 == 0:
            cell.fill = ALT_FILL
        cell.alignment = Alignment(
            horizontal='left' if ci == 1 else 'center',
            vertical='center', wrap_text=(ci == 1))

ws.row_dimensions[header_row].height = 30
ws.column_dimensions['A'].width = 38
for i in range(2, len(cohorts) + 2):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

ws.freeze_panes = f'B{header_row + 1}'
wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
