"""
stroke_landmark_table.py

Outputs the 30-day landmark SLP timing analysis results to Excel.

Sheets:
  1. LM1_Cohort       — N, baseline characteristics by timing group
  2. LM2_Balance      — Covariate SMDs across timing groups
  3. LM3_CoxPH        — Adjusted HRs, 95% CI, p-value, E-value
  4. LM4_KM           — Cumulative incidence % at 90/180/365/730d by group
  5. LM5_Sensitivity  — Overlap trimming sensitivity (mortality)
  6. LM6_Elix         — Cox PH stratified by Elixhauser (van Walraven) quartile
  7. LM7_ByDischDest  — Cox PH stratified by discharge destination (home/SNF/IRF/HHA)
"""

import sys
sys.path.insert(0, r'C:\users\hsaee\desktop\cms_viewer\env\Lib\site-packages')

import duckdb
import numpy as np
import pandas as pd
from lifelines import CoxPHFitter, KaplanMeierFitter
from lifelines.statistics import logrank_test
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

DB_PATH  = r"F:\CMS\cms_data.duckdb"
OUT_PATH = r"F:\CMS\projects\stroke_SLP\stroke_landmark_tables_home_hha.xlsx"
LANDMARK = 90
MAX_FOLLOW = 1885
MAX_FOLLOW_LM = MAX_FOLLOW - LANDMARK

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data...")
con = duckdb.connect(DB_PATH, read_only=True)
con.execute("SET memory_limit='24GB'; SET threads=12;")
df = con.execute("""
    SELECT
        p.DSYSRTKY, p.age_at_adm, p.sex, p.stroke_type,
        p.index_los, p.mech_vent, p.peg_placed, p.trach_placed,
        p.dschg_status, p.dysphagia_poa, p.aspiration_poa,
        p.prior_stroke, p.dementia, p.van_walraven_score, p.prop_score,
        p.slp_outpt_any_90d, p.days_to_slp_outpt,
        o.days_to_death, o.days_to_aspiration, o.has_aspiration,
        o.days_to_gtube, o.has_gtube, o.pre_stroke_tube,
        o.days_to_dysphagia, o.has_dysphagia,
        o.days_to_recur_stroke, o.days_to_readmit
    FROM stroke_propensity p
    JOIN stroke_outcomes o ON o.DSYSRTKY = p.DSYSRTKY
""").df()
con.close()
print(f"  Loaded {len(df):,} rows")

# ── Landmark restriction ───────────────────────────────────────────────────────
lm = df[df['days_to_death'].isna() | (df['days_to_death'] > LANDMARK)].copy()
lm_slp = lm[lm['slp_outpt_any_90d'] == 1].copy()

def slp_group(row):
    d = row['days_to_slp_outpt']
    if pd.isna(d): return None
    elif d <= 14:  return 'SLP 0-14d'
    elif d <= 30:  return 'SLP 15-30d'
    else:          return 'SLP 31-90d'

lm_slp['timing_group'] = lm_slp.apply(slp_group, axis=1)

# ── Shift outcome times ────────────────────────────────────────────────────────
for col in ['days_to_death', 'days_to_aspiration', 'days_to_gtube',
            'days_to_dysphagia', 'days_to_recur_stroke', 'days_to_readmit']:
    lm_slp[col] = lm_slp[col].astype('float64')
    lm_slp[f'{col}_lm'] = np.where(
        lm_slp[col].notna() & (lm_slp[col] > LANDMARK),
        (lm_slp[col] - LANDMARK).clip(lower=0.5), np.nan
    )

lm_slp['censor_lm'] = np.where(
    lm_slp['days_to_death_lm'].notna(),
    lm_slp['days_to_death_lm'].clip(upper=MAX_FOLLOW_LM), MAX_FOLLOW_LM
)

lm_slp['died_lm']      = lm_slp['days_to_death_lm'].notna().astype(int)
lm_slp['asp_lm']       = lm_slp['days_to_aspiration_lm'].notna().astype(int)
lm_slp['gtube_lm']     = lm_slp['days_to_gtube_lm'].notna().astype(int)
lm_slp['recur_lm']     = lm_slp['days_to_recur_stroke_lm'].notna().astype(int)

for ev, dur in [('asp_lm',   'days_to_aspiration_lm'),
                ('gtube_lm', 'days_to_gtube_lm'),
                ('recur_lm', 'days_to_recur_stroke_lm')]:
    lm_slp[f'{dur}_filled'] = lm_slp[dur].fillna(lm_slp['censor_lm']).clip(lower=0.5)

# ── Covariates ─────────────────────────────────────────────────────────────────
lm_slp['slp_15_30d'] = (lm_slp['timing_group'] == 'SLP 15-30d').astype(int)
lm_slp['slp_31_90d'] = (lm_slp['timing_group'] == 'SLP 31-90d').astype(int)

def map_dschg(code):
    if pd.isna(code): return 'home'
    c = str(code).strip()
    if c in ('01', '08'): return 'home'
    if c == '06':         return 'hha'
    if c in ('03', '61'): return 'snf'
    if c == '62':         return 'irf'
    return 'other'

lm_slp['dschg_group'] = lm_slp['dschg_status'].apply(map_dschg)

# Restrict to home and HHA discharge only
lm_slp = lm_slp[lm_slp['dschg_group'].isin(['home', 'hha'])].copy()
print(f"  After home/HHA filter: {len(lm_slp):,} rows")

dschg_dummies  = pd.get_dummies(lm_slp['dschg_group'], prefix='dschg', drop_first=True)
dschg_dummies  = dschg_dummies.loc[:, dschg_dummies.std() > 0.01]
sex_dummies    = pd.get_dummies(lm_slp['sex'],         prefix='sex',    drop_first=True)
stroke_dummies = pd.get_dummies(lm_slp['stroke_type'], prefix='stroke', drop_first=True)
lm_slp = pd.concat([lm_slp, dschg_dummies, sex_dummies, stroke_dummies], axis=1)

_cov_candidates = (
    ['slp_15_30d', 'slp_31_90d',
     'age_at_adm', 'van_walraven_score', 'index_los',
     'mech_vent', 'trach_placed', 'prior_stroke', 'dementia',
     'dysphagia_poa', 'aspiration_poa']
    + list(dschg_dummies.columns)
    + list(sex_dummies.columns)
    + list(stroke_dummies.columns)
)
# Drop near-zero-variance covariates (e.g. trach_placed is ~0 in home/HHA cohort)
# Threshold 0.05 drops any binary variable with <0.25% prevalence
COX_COVARIATES = [c for c in _cov_candidates if lm_slp[c].std() > 0.05]
_dropped = set(_cov_candidates) - set(COX_COVARIATES)
if _dropped:
    print(f"  Dropped near-zero-variance covariates: {_dropped}")

GROUPS   = ['SLP 0-14d', 'SLP 15-30d', 'SLP 31-90d']
TIME_PTS = [90, 180, 365, 730]

# ── Helper: SMD ────────────────────────────────────────────────────────────────
def smd(x1, x0):
    m1, m0 = x1.mean(), x0.mean()
    s1, s0 = x1.std(), x0.std()
    ps = np.sqrt((s1**2 + s0**2) / 2)
    return abs(m1 - m0) / ps if ps > 0 else 0.0

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Cohort characteristics
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM1_Cohort...")

def cohort_row(label, vals_by_grp, fmt='{:.1f}'):
    row = {'Characteristic': label}
    for g, v in zip(GROUPS, vals_by_grp):
        row[g] = fmt.format(v) if v is not None else ''
    return row

rows = []
for grp in GROUPS:
    g = lm_slp[lm_slp['timing_group'] == grp]
    rows.append({'Characteristic': f'N', **{gr: '' for gr in GROUPS}})
    break

# Build properly
records = []
grp_dfs = {g: lm_slp[lm_slp['timing_group'] == g] for g in GROUPS}
n_vals  = [len(grp_dfs[g]) for g in GROUPS]

records.append({'Characteristic': 'N', **{g: f"{len(grp_dfs[g]):,}" for g in GROUPS}})
records.append({'Characteristic': 'Age, mean (SD)',
    **{g: f"{grp_dfs[g]['age_at_adm'].mean():.1f} ({grp_dfs[g]['age_at_adm'].std():.1f})"
       for g in GROUPS}})
records.append({'Characteristic': 'Female, %',
    **{g: f"{100*(grp_dfs[g]['sex']=='Female').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'White race, %',
    **{g: '' for g in GROUPS}})   # not in propensity — placeholder
records.append({'Characteristic': 'Ischemic stroke, %',
    **{g: f"{100*(grp_dfs[g]['stroke_type']=='Ischemic').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'ICH, %',
    **{g: f"{100*(grp_dfs[g]['stroke_type']=='ICH').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'SAH, %',
    **{g: f"{100*(grp_dfs[g]['stroke_type']=='SAH').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Index LOS, mean (SD)',
    **{g: f"{grp_dfs[g]['index_los'].mean():.1f} ({grp_dfs[g]['index_los'].std():.1f})"
       for g in GROUPS}})
records.append({'Characteristic': 'van Walraven score, mean (SD)',
    **{g: f"{grp_dfs[g]['van_walraven_score'].mean():.1f} ({grp_dfs[g]['van_walraven_score'].std():.1f})"
       for g in GROUPS}})
records.append({'Characteristic': 'Dysphagia POA, %',
    **{g: f"{100*grp_dfs[g]['dysphagia_poa'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Aspiration POA, %',
    **{g: f"{100*grp_dfs[g]['aspiration_poa'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Mechanical ventilation, %',
    **{g: f"{100*grp_dfs[g]['mech_vent'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Tracheostomy, %',
    **{g: f"{100*grp_dfs[g]['trach_placed'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'PEG at index admission, %',
    **{g: f"{100*grp_dfs[g]['peg_placed'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Prior stroke, %',
    **{g: f"{100*grp_dfs[g]['prior_stroke'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Dementia, %',
    **{g: f"{100*grp_dfs[g]['dementia'].mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Discharge to IRF, %',
    **{g: f"{100*(grp_dfs[g]['dschg_group']=='irf').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Discharge to home, %',
    **{g: f"{100*(grp_dfs[g]['dschg_group']=='home').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Discharge to SNF, %',
    **{g: f"{100*(grp_dfs[g]['dschg_group']=='snf').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Discharge to HHA, %',
    **{g: f"{100*(grp_dfs[g]['dschg_group']=='hha').mean():.1f}" for g in GROUPS}})
records.append({'Characteristic': 'Days to first SLP, median [IQR]',
    **{g: f"{grp_dfs[g]['days_to_slp_outpt'].median():.0f} "
          f"[{grp_dfs[g]['days_to_slp_outpt'].quantile(0.25):.0f}–"
          f"{grp_dfs[g]['days_to_slp_outpt'].quantile(0.75):.0f}]"
       for g in GROUPS}})

df_cohort = pd.DataFrame(records)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Covariate balance (SMDs)
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM2_Balance...")

cont_vars = ['age_at_adm', 'index_los', 'van_walraven_score', 'days_to_slp_outpt']
bin_vars  = ['dysphagia_poa', 'aspiration_poa', 'mech_vent', 'trach_placed',
             'peg_placed', 'prior_stroke', 'dementia']

bal_records = []
ref = grp_dfs['SLP 0-14d']
for var in cont_vars + bin_vars:
    row = {'Variable': var}
    for g in ['SLP 15-30d', 'SLP 31-90d']:
        row[f'SMD vs 0-14d ({g})'] = round(smd(grp_dfs[g][var].dropna(),
                                               ref[var].dropna()), 3)
    bal_records.append(row)

df_balance = pd.DataFrame(bal_records)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Cox PH results
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM3_CoxPH...")

cox_models = [
    ('Aspiration PNA (PRIMARY)',  'asp_lm',   'days_to_aspiration_lm_filled',   False, 0.0),
    ('PEG/G-tube (PRIMARY)',      'gtube_lm', 'days_to_gtube_lm_filled',        True,  0.1),
    ('Mortality (secondary)',     'died_lm',  'censor_lm',                      False, 0.0),
    ('Recurrent stroke (sec)',    'recur_lm', 'days_to_recur_stroke_lm_filled', False, 0.0),
]

cox_records = []
for label, ev_col, dur_col, excl_peg, pen in cox_models:
    sub = lm_slp.copy()
    if excl_peg:
        sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
    cov_use = list(COX_COVARIATES)
    if not excl_peg and 'peg_placed' not in cov_use:
        cov_use = ['peg_placed'] + cov_use

    cox_df = sub[[dur_col, ev_col] + cov_use].rename(
        columns={dur_col: 'duration', ev_col: 'event'}
    ).dropna()

    n_ev = int(cox_df['event'].sum())
    if n_ev < 20:
        continue

    try:
        cph = CoxPHFitter(penalizer=pen)
        cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
        for tv in ['slp_15_30d', 'slp_31_90d']:
            if tv not in cph.summary.index:
                continue
            r = cph.summary.loc[tv]
            hr   = np.exp(r['coef'])
            lo95 = np.exp(r['coef lower 95%'])
            hi95 = np.exp(r['coef upper 95%'])
            p    = r['p']
            ev_val = hr + np.sqrt(hr * abs(hr - 1)) if not np.isnan(hr) else np.nan
            cox_records.append({
                'Outcome':     label,
                'Comparison':  '15-30d vs 0-14d' if tv == 'slp_15_30d' else '31-90d vs 0-14d',
                'N':           len(cox_df),
                'Events':      n_ev,
                'Event %':     round(100 * n_ev / len(cox_df), 1),
                'HR':          round(hr, 2),
                'CI_low':      round(lo95, 2),
                'CI_high':     round(hi95, 2),
                '95% CI':      f"[{lo95:.2f}, {hi95:.2f}]",
                'p-value':     '<0.0001' if p < 0.0001 else f'{p:.4f}',
                'E-value':     round(ev_val, 2),
                'Excl index PEG': 'Yes' if excl_peg else 'No',
            })
    except Exception as e:
        cox_records.append({
            'Outcome': label, 'Comparison': 'ERROR', 'N': 0, 'Events': 0,
            'Event %': 0, 'HR': None, 'CI_low': None, 'CI_high': None,
            '95% CI': str(e)[:60], 'p-value': '', 'E-value': None,
            'Excl index PEG': ''
        })

df_cox = pd.DataFrame(cox_records)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: KM cumulative incidence
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM4_KM...")

km_outcomes = [
    ('Aspiration PNA',  'asp_lm',    'days_to_aspiration_lm_filled', False),
    ('PEG/G-tube',      'gtube_lm',  'days_to_gtube_lm_filled',      True),
    ('Mortality',       'died_lm',   'censor_lm',                    False),
    ('Recurrent Stroke','recur_lm',  'days_to_recur_stroke_lm_filled',False),
]

km_records = []
for label, ev_col, dur_col, excl_peg in km_outcomes:
    sub = lm_slp.copy()
    if excl_peg:
        sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]

    grps_sub = {g: sub[sub['timing_group'] == g] for g in GROUPS}
    kmfs = {}
    for g, gdf in grps_sub.items():
        kmf = KaplanMeierFitter()
        valid = gdf[dur_col].notna()
        kmf.fit(gdf.loc[valid, dur_col], gdf.loc[valid, ev_col], label=g)
        kmfs[g] = kmf

    # Log-rank 0-14d vs 31-90d
    g0 = grps_sub['SLP 0-14d'];  g2 = grps_sub['SLP 31-90d']
    v0 = g0[dur_col].notna();    v2 = g2[dur_col].notna()
    lr = logrank_test(g0.loc[v0, dur_col], g2.loc[v2, dur_col],
                      event_observed_A=g0.loc[v0, ev_col],
                      event_observed_B=g2.loc[v2, ev_col])
    p_lr = '<0.0001' if lr.p_value < 0.0001 else f'{lr.p_value:.4f}'

    for t in TIME_PTS:
        row = {'Outcome': label, 'Days from landmark': t,
               'Log-rank p (0-14d vs 31-90d)': p_lr}
        for g in GROUPS:
            kmf = kmfs[g]
            s   = float(kmf.predict(t))
            ci  = kmf.confidence_interval_survival_function_
            idx = min(ci.index.searchsorted(t, side='right'), len(ci) - 1)
            lo  = float(ci.iloc[idx, 0])
            hi  = float(ci.iloc[idx, 1])
            pct      = round(100 * (1 - s),  1)
            pct_lo   = round(100 * (1 - hi), 1)
            pct_hi   = round(100 * (1 - lo), 1)
            row[f'{g} — % (95% CI)'] = f"{pct:.1f} [{pct_lo:.1f}, {pct_hi:.1f}]"
            row[f'{g} — % (raw)']    = pct
        km_records.append(row)

df_km = pd.DataFrame(km_records)
# Display columns: drop raw columns for main view
km_display_cols = (['Outcome', 'Days from landmark', 'Log-rank p (0-14d vs 31-90d)'] +
                   [f'{g} — % (95% CI)' for g in GROUPS])
df_km_display = df_km[km_display_cols]

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Overlap trimming sensitivity
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM5_Sensitivity...")

sens_records = []
n_before = len(lm_slp)
lm_trim  = lm_slp[(lm_slp['prop_score'] >= 0.05) & (lm_slp['prop_score'] <= 0.95)].copy()
n_after  = len(lm_trim)

for dataset, label in [(lm_slp, 'Full cohort'), (lm_trim, f'PS trimmed (0.05–0.95)')]:
    for out_label, ev_col, dur_col, excl_peg, pen in [
        ('Aspiration PNA', 'asp_lm',   'days_to_aspiration_lm_filled', False, 0.0),
        ('PEG/G-tube',     'gtube_lm', 'days_to_gtube_lm_filled',      True,  0.1),
        ('Mortality',      'died_lm',  'censor_lm',                    False, 0.0),
    ]:
        sub = dataset.copy()
        if excl_peg:
            sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
        cov_use = list(COX_COVARIATES)
        if not excl_peg and 'peg_placed' not in cov_use:
            cov_use = ['peg_placed'] + cov_use

        cox_df = sub[[dur_col, ev_col] + cov_use].rename(
            columns={dur_col: 'duration', ev_col: 'event'}
        ).dropna()

        try:
            cph = CoxPHFitter(penalizer=pen)
            cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
            for tv in ['slp_15_30d', 'slp_31_90d']:
                if tv not in cph.summary.index:
                    continue
                r    = cph.summary.loc[tv]
                hr   = np.exp(r['coef'])
                lo95 = np.exp(r['coef lower 95%'])
                hi95 = np.exp(r['coef upper 95%'])
                p    = r['p']
                sens_records.append({
                    'Dataset':    label,
                    'N':          len(cox_df),
                    'N excluded': n_before - n_after if label != 'Full cohort' else 0,
                    'Outcome':    out_label,
                    'Comparison': '15-30d vs 0-14d' if tv == 'slp_15_30d' else '31-90d vs 0-14d',
                    'HR':         round(hr, 2),
                    '95% CI':     f"[{lo95:.2f}, {hi95:.2f}]",
                    'p-value':    '<0.0001' if p < 0.0001 else f'{p:.4f}',
                })
        except Exception as e:
            sens_records.append({
                'Dataset': label, 'N': 0, 'N excluded': 0,
                'Outcome': out_label, 'Comparison': 'ERROR',
                'HR': None, '95% CI': str(e)[:60], 'p-value': ''
            })

df_sens = pd.DataFrame(sens_records)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: Cox PH stratified by Elixhauser (van Walraven) quartile
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM6_Elix...")

vw = lm_slp['van_walraven_score']
q1, q2, q3 = vw.quantile([0.25, 0.50, 0.75])

ql = [
    f'Q1 (≤{q1:.0f})',
    f'Q2 ({q1:.0f}–{q2:.0f})',
    f'Q3 ({q2:.0f}–{q3:.0f})',
    f'Q4 (>{q3:.0f})',
]

def vw_group(score):
    if pd.isna(score): return None
    if score <= q1: return ql[0]
    if score <= q2: return ql[1]
    if score <= q3: return ql[2]
    return ql[3]

lm_slp['vw_quartile'] = lm_slp['van_walraven_score'].apply(vw_group)

elix_records = []
for qlab in ql:
    q_sub = lm_slp[lm_slp['vw_quartile'] == qlab]
    n_q   = len(q_sub)
    vw_range = (q_sub['van_walraven_score'].min(), q_sub['van_walraven_score'].max())

    for label, ev_col, dur_col, excl_peg, pen in cox_models:
        sub = q_sub.copy()
        if excl_peg:
            sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
        cov_use = list(COX_COVARIATES)
        if not excl_peg and 'peg_placed' not in cov_use:
            cov_use = ['peg_placed'] + cov_use

        cox_df = sub[[dur_col, ev_col] + cov_use].rename(
            columns={dur_col: 'duration', ev_col: 'event'}
        ).dropna()

        n_ev = int(cox_df['event'].sum())

        for tv, comp in [('slp_15_30d', '15-30d vs 0-14d'),
                         ('slp_31_90d', '31-90d vs 0-14d')]:
            base = {
                'Elixhauser Quartile': qlab,
                'VW range':            f"{vw_range[0]:.0f} to {vw_range[1]:.0f}",
                'N in quartile':       n_q,
                'Outcome':             label,
                'Comparison':          comp,
                'N':                   len(cox_df),
                'Events':              n_ev,
            }
            if n_ev < 10:
                elix_records.append({**base, 'HR': None, '95% CI': 'Too few events', 'p-value': ''})
                continue
            try:
                cph = CoxPHFitter(penalizer=pen)
                cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
                if tv not in cph.summary.index:
                    elix_records.append({**base, 'HR': None, '95% CI': 'Not estimable', 'p-value': ''})
                    continue
                r    = cph.summary.loc[tv]
                hr   = np.exp(r['coef'])
                lo95 = np.exp(r['coef lower 95%'])
                hi95 = np.exp(r['coef upper 95%'])
                p    = r['p']
                elix_records.append({
                    **base,
                    'HR':      round(hr, 2),
                    '95% CI':  f"[{lo95:.2f}, {hi95:.2f}]",
                    'p-value': '<0.0001' if p < 0.0001 else f'{p:.4f}',
                })
            except Exception as e:
                elix_records.append({**base, 'HR': None, '95% CI': str(e)[:60], 'p-value': ''})

df_elix = pd.DataFrame(elix_records)

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Cox PH stratified by discharge destination
# ═══════════════════════════════════════════════════════════════════════════════
print("Building LM7_ByDischDest...")

# Within each discharge destination stratum, discharge dummies are not needed
# (destination is fixed); all other covariates retained.
COX_COVARIATES_NODSCHG = (
    ['slp_15_30d', 'slp_31_90d',
     'age_at_adm', 'van_walraven_score', 'index_los',
     'mech_vent', 'trach_placed', 'prior_stroke', 'dementia',
     'dysphagia_poa', 'aspiration_poa']
    + list(sex_dummies.columns)
    + list(stroke_dummies.columns)
)

DSCHG_STRATA = [
    ('home', 'Home / self-care (01, 08)'),
    ('hha',  'Home Health Agency (06)'),
]

dest_records = []
for dschg_code, dschg_label in DSCHG_STRATA:
    d_sub = lm_slp[lm_slp['dschg_group'] == dschg_code]
    n_d   = len(d_sub)

    # Group sizes within stratum
    n_by_grp = {g: int((d_sub['timing_group'] == g).sum()) for g in GROUPS}

    for label, ev_col, dur_col, excl_peg, pen in cox_models:
        sub = d_sub.copy()
        if excl_peg:
            sub = sub[(sub['peg_placed'] == 0) & (sub['pre_stroke_tube'].fillna(0) == 0)]
        cov_use = list(COX_COVARIATES_NODSCHG)
        if not excl_peg and 'peg_placed' not in cov_use:
            cov_use = ['peg_placed'] + cov_use

        # Rename duration/event, then drop constant columns within stratum
        raw = sub[[dur_col, ev_col] + cov_use].rename(
            columns={dur_col: 'duration', ev_col: 'event'}
        ).dropna()
        cov_use_ok = [c for c in cov_use if raw[c].std() > 0.01]
        cox_df = raw[['duration', 'event'] + cov_use_ok]

        n_ev = int(cox_df['event'].sum())

        for tv, comp in [('slp_15_30d', '15-30d vs 0-14d'),
                         ('slp_31_90d', '31-90d vs 0-14d')]:
            base = {
                'Discharge destination': dschg_label,
                'N in stratum':          n_d,
                'N (0-14d)':             n_by_grp['SLP 0-14d'],
                'N (15-30d)':            n_by_grp['SLP 15-30d'],
                'N (31-90d)':            n_by_grp['SLP 31-90d'],
                'Outcome':               label,
                'Comparison':            comp,
                'N in model':            len(cox_df),
                'Events':                n_ev,
            }
            if (n_ev < 20
                    or tv not in cox_df.columns
                    or (tv == 'slp_15_30d' and n_by_grp['SLP 15-30d'] < 50)
                    or (tv == 'slp_31_90d' and n_by_grp['SLP 31-90d'] < 50)):
                dest_records.append({**base, 'HR': None, '95% CI': 'Too few', 'p-value': ''})
                continue
            try:
                cph = CoxPHFitter(penalizer=pen)
                cph.fit(cox_df, duration_col='duration', event_col='event', show_progress=False)
                if tv not in cph.summary.index:
                    dest_records.append({**base, 'HR': None, '95% CI': 'Not estimable', 'p-value': ''})
                    continue
                r    = cph.summary.loc[tv]
                hr   = np.exp(r['coef'])
                lo95 = np.exp(r['coef lower 95%'])
                hi95 = np.exp(r['coef upper 95%'])
                p    = r['p']
                ev_val = hr + np.sqrt(hr * abs(hr - 1)) if not np.isnan(hr) else np.nan
                dest_records.append({
                    **base,
                    'HR':      round(hr, 2),
                    '95% CI':  f"[{lo95:.2f}, {hi95:.2f}]",
                    'p-value': '<0.0001' if p < 0.0001 else f'{p:.4f}',
                    'E-value': round(ev_val, 2),
                })
            except Exception as e:
                dest_records.append({**base, 'HR': None, '95% CI': str(e)[:60], 'p-value': ''})

df_dest = pd.DataFrame(dest_records)

# ═══════════════════════════════════════════════════════════════════════════════
# Write Excel
# ═══════════════════════════════════════════════════════════════════════════════
print(f"Writing {OUT_PATH} ...")

HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', size=10)
ALT_FILL     = PatternFill('solid', fgColor='D6E4F0')
BORDER_THIN  = Border(
    bottom=Side(style='thin', color='B0B0B0'),
    right =Side(style='thin', color='B0B0B0'),
)
TITLE_FONT   = Font(bold=True, size=12, color='1F4E79')

def write_sheet(wb, sheet_name, df, title):
    ws = wb.create_sheet(title=sheet_name)
    ws.append([title])
    ws['A1'].font = TITLE_FONT
    ws.append([])

    header_row = ws.max_row + 1
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font    = HEADER_FONT
        cell.fill    = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for r_idx, row in enumerate(df.itertuples(index=False), header_row + 1):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = Alignment(horizontal='left')
            if fill:
                cell.fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    return ws

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

write_sheet(wb, 'LM1_Cohort',
    df_cohort,
    f'Table LM1: Baseline Characteristics by SLP Timing Group '
    f'(N={len(lm_slp):,} outpatient SLP recipients within 90d of discharge; '
    f'90-day landmark cohort)')

write_sheet(wb, 'LM2_Balance',
    df_balance,
    'Table LM2: Covariate Balance Across SLP Timing Groups (SMD vs 0–14d reference)')

write_sheet(wb, 'LM3_CoxPH',
    df_cox[['Outcome','Comparison','N','Events','Event %','HR','95% CI','p-value','E-value','Excl index PEG']],
    'Table LM3: Adjusted Cox PH — Outcomes from 90-Day Landmark '
    '(Reference: SLP 0–14d; adjusted for age, comorbidity, LOS, mech vent, trach, '
    'prior stroke, dementia, dysphagia POA, aspiration POA, discharge destination, sex, stroke type)')

write_sheet(wb, 'LM4_KM',
    df_km_display,
    'Table LM4: Kaplan–Meier Cumulative Incidence by SLP Timing Group '
    '(% from 90-day landmark [95% CI]; log-rank 0–14d vs 31–90d)')

write_sheet(wb, 'LM5_Sensitivity',
    df_sens,
    f'Table LM5: Overlap Trimming Sensitivity Analysis '
    f'(Full N={n_before:,} vs PS-trimmed N={n_after:,}; '
    f'excluded {n_before-n_after:,} with PS <0.05 or >0.95)')

write_sheet(wb, 'LM6_Elix',
    df_elix,
    f'Table LM6: Adjusted Cox PH Stratified by Elixhauser Comorbidity (van Walraven) Quartile '
    f'(Reference: SLP 0–14d; cutpoints Q1≤{q1:.0f}, Q2≤{q2:.0f}, Q3≤{q3:.0f})')

write_sheet(wb, 'LM7_ByDischDest',
    df_dest,
    'Table LM7: Adjusted Cox PH Stratified by Discharge Destination '
    '(Reference: SLP 0–14d; discharge destination fixed within stratum — '
    'addresses IRF/SNF timing confounding)')

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Sheets: {', '.join(ws.title for ws in wb.worksheets)}")
